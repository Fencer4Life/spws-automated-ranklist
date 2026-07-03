"""
Plan-test-ID 5.M9 (ADR-066): PPW1 XLSX parser handles single-competitor
walkover sheets.

SPWS PPW1 ranking sheets ("PP1") follow a Polish-localized template with
"Miejsce" (Place) and "Nazwisko Imię" (Name) columns. When a V-cat has
only one registrant, the organizer leaves the Place column empty (no
ranking to compute). The parser must still emit one ParsedResult with
place=1 so the threshold gate (ADR-066) can decide whether to ingest.
"""

from __future__ import annotations

import tempfile
from pathlib import Path


def _make_pp1_xlsx(rows_after_header):
    """Build a tempfile XLSX with a `PP1` sheet matching the SPWS template.

    `rows_after_header` is a list of (place_or_None, name, country_or_None)
    tuples. Header row is auto-generated.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    # Replace default sheet with PP1
    assert wb.active is not None  # a freshly created Workbook() always has an active sheet
    wb.remove(wb.active)
    ws = wb.create_sheet("PP1")
    ws.append(["Miejsce\nPlace", "Nazwisko Imię\nName", "Państwo\nCountry"])
    for place, name, country in rows_after_header:
        ws.append([place, name, country])
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)
    return Path(tmp.name)


def test_5_M9_1_single_fencer_no_place_gets_walkover_place_1():
    """5.M9.1 — PP1 sheet with one fencer-row and empty Place → place=1."""
    from python.tools.recreate_active_season_2025_2026 import _parse_ppw1_xlsx

    p = _make_pp1_xlsx([(None, "KOWALSKI Jan", "POL")])
    parsed = _parse_ppw1_xlsx(p)
    assert len(parsed.results) == 1
    assert parsed.results[0].place == 1
    assert parsed.results[0].fencer_name == "KOWALSKI Jan"


def test_5_M9_2_two_fencers_no_places_drops_both():
    """5.M9.2 — Walkover does NOT activate for n=2 with all-null places.
    These rows are dropped (no valid placements to score)."""
    from python.tools.recreate_active_season_2025_2026 import _parse_ppw1_xlsx

    p = _make_pp1_xlsx(
        [
            (None, "A", "POL"),
            (None, "B", "POL"),
        ]
    )
    parsed = _parse_ppw1_xlsx(p)
    assert len(parsed.results) == 0


def test_5_M9_3_empty_pp1_yields_no_results():
    """5.M9.3 — header-only sheet (zero data rows) → zero results;
    walkover requires at least one named fencer."""
    from python.tools.recreate_active_season_2025_2026 import _parse_ppw1_xlsx

    p = _make_pp1_xlsx([])
    parsed = _parse_ppw1_xlsx(p)
    assert len(parsed.results) == 0


def test_5_M9_4_single_fencer_with_explicit_place_keeps_it():
    """5.M9.4 — explicit place=1 wins (walkover is a fallback, not
    an override)."""
    from python.tools.recreate_active_season_2025_2026 import _parse_ppw1_xlsx

    p = _make_pp1_xlsx([(1, "WIŚNIEWSKI Piotr", "POL")])
    parsed = _parse_ppw1_xlsx(p)
    assert len(parsed.results) == 1
    assert parsed.results[0].place == 1


def test_5_M9_5_three_fencers_with_places_no_walkover_needed():
    """5.M9.5 — normal multi-fencer sheet stays as today."""
    from python.tools.recreate_active_season_2025_2026 import _parse_ppw1_xlsx

    p = _make_pp1_xlsx(
        [
            (1, "A", "POL"),
            (2, "B", "POL"),
            (3, "C", "POL"),
        ]
    )
    parsed = _parse_ppw1_xlsx(p)
    assert len(parsed.results) == 3
    assert [r.place for r in parsed.results] == [1, 2, 3]
