"""
generate_season_seed.py
-----------------------
Reads an SPWS season Excel workbook and generates a seed SQL file for one
age category, written to supabase/data/{season_folder}/{age_cat}_{gender}_{weapon}.sql.

Usage:
    python python/tools/generate_season_seed.py \\
        --xlsx reference/SZPADA-2-2024-2025.xlsx \\
        --season SPWS-2024-2025 \\
        --weapon EPEE \\
        --gender M \\
        --age-cat V2

Output path is derived automatically:
    season SPWS-2024-2025 + weapon EPEE + gender M + age-cat V2
    → supabase/data/2024_25/v2_m_epee.sql

The file is auto-loaded by `supabase db reset` via config.toml sql_paths glob.
"""

import argparse
import re
import sys
import math
from pathlib import Path
from datetime import date, datetime

import openpyxl
import psycopg2
from rapidfuzz import fuzz

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DB_URL = "postgresql://postgres:postgres@127.0.0.1:54322/postgres"

# Map sheet name → (tournament_type, event_code_prefix, human_name)
SHEET_MAP = {
    "PP1":  ("PPW", "PP1",  "I Puchar Polski Weteranów"),
    "PP2":  ("PPW", "PP2",  "II Puchar Polski Weteranów"),
    "PP3":  ("PPW", "PP3",  "III Puchar Polski Weteranów"),
    "PP4":  ("PPW", "PP4",  "IV Puchar Polski Weteranów"),
    "PP5":  ("PPW", "PP5",  "V Puchar Polski Weteranów"),
    "GP7":  ("PPW", "GP7",  "Grand Prix (runda 7)"),
    "GP8":  ("PPW", "GP8",  "Grand Prix (runda 8)"),
    "MPW":  ("MPW", "MPW",  "Mistrzostwa Polski Weteranów"),
    "PEW1": ("PEW", "PEW1", "EVF Grand Prix 1 — Budapeszt"),
    "PEW2": ("PEW", "PEW2", "EVF Grand Prix 2 — Madryt"),
    "PEW3": ("PEW", "PEW3", "EVF Grand Prix 3"),
    "PEW4": ("PEW", "PEW4", "EVF Grand Prix 4"),
    "PEW5": ("PEW", "PEW5", "EVF Grand Prix 5"),
    "PEW6": ("PEW", "PEW6", "EVF Grand Prix 6"),
    "PEW7": ("PEW", "PEW7", "EVF Grand Prix 7 — Terni"),
    "PEW8": ("PEW", "PEW8", "EVF Grand Prix 8 — Guildford"),
    "PEW9": ("PEW", "PEW9", "EVF Grand Prix 9 — Sztokholm"),
    "PEW10":("PEW", "PEW10","EVF Grand Prix 10 — Graz"),
    "PEW11":("PEW", "PEW11","EVF Grand Prix 11 — Gdańsk"),
    "PEW12":("PEW", "PEW12","EVF Grand Prix 12 — Ateny"),
    "PS":   ("PSW", "PS",   "Puchar Świata"),
    "IMEW": ("MEW", "IMEW", "Indywidualne Mistrzostwa Europy Weteranów"),
}

MATCH_THRESHOLD = 80  # minimum fuzzy score to accept a match

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sq(s):
    """Escape single quotes for SQL."""
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def parse_date(raw) -> str | None:
    """Return ISO date string from various date formats in the Excel."""
    if raw is None:
        return None
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    # "DD.MM.YYYY"
    m = re.match(r"(\d{1,2})\.(\d{2})\.(\d{4})", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    # "D-D.MM.YYYY" multi-day event
    m = re.match(r"(\d{1,2})-\d{1,2}\.(\d{2})\.(\d{4})", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    # bare year
    m = re.match(r"^(\d{4})$", s)
    if m:
        return f"{s}-01-01"
    return None


def normalize_name(name: str) -> str:
    """Uppercase + strip extra whitespace."""
    return " ".join(name.upper().split())


def fuzzy_match(name: str, fencers: list[tuple]) -> tuple | None:
    """
    Return (id_fencer, fencer_name, score) or None.
    fencers: list of (id, txt_surname, txt_first_name, aliases)
    Checks aliases first (exact, case-insensitive), then fuzzy on full name.
    """
    norm = normalize_name(name)
    # 1. Alias exact match
    for fid, surname, first, aliases in fencers:
        for alias in aliases:
            if normalize_name(alias) == norm:
                return (fid, f"{surname} {first}", 100)
    # 2. Fuzzy match on full name
    best_score = 0
    best_fencer = None
    for fid, surname, first, aliases in fencers:
        full = f"{surname} {first}"
        score = fuzz.token_sort_ratio(norm, full.upper())
        if score > best_score:
            best_score = score
            best_fencer = (fid, f"{surname} {first}")
    if best_score >= MATCH_THRESHOLD:
        return (*best_fencer, best_score)
    return None


def load_fencers(conn) -> list[tuple]:
    """Return list of (id, surname, first_name, aliases_list)."""
    cur = conn.cursor()
    cur.execute("""
        SELECT id_fencer, txt_surname, txt_first_name,
               COALESCE(json_name_aliases::text, '[]')
          FROM tbl_fencer
         ORDER BY id_fencer
    """)
    rows = []
    for fid, surname, first, aliases_json in cur.fetchall():
        import json
        aliases = json.loads(aliases_json)
        rows.append((fid, surname, first, aliases))
    return rows


# ---------------------------------------------------------------------------
# Extract tournament data from one Excel sheet
# ---------------------------------------------------------------------------

def extract_sheet(wb_data, wb_links, sheet_name: str) -> dict:
    """Return dict with tournament metadata and result rows."""
    ws_d = wb_data[sheet_name]
    ws_l = wb_links[sheet_name]

    location = ws_d.cell(2, 3).value
    date_raw  = ws_d.cell(3, 3).value
    n         = ws_d.cell(2, 8).value   # participant count

    url = None
    hl  = ws_l.cell(2, 3).hyperlink
    if hl:
        url = hl.target

    results = []
    for r in range(6, ws_d.max_row + 1):
        name  = ws_d.cell(r, 3).value
        place = ws_d.cell(r, 8).value
        if not name or name in ("x", "X", ""):
            continue
        if not isinstance(place, int):
            continue
        results.append({"name": str(name).strip(), "place": place})

    return {
        "location": location,
        "date":     parse_date(date_raw),
        "n":        n if isinstance(n, int) else None,
        "url":      url,
        "results":  results,
    }


# ---------------------------------------------------------------------------
# Derive output path and SQL identifiers from CLI args
# ---------------------------------------------------------------------------

def derive_paths(season: str, weapon: str, gender: str, age_cat: str) -> tuple[Path, str, str]:
    """
    Return (out_path, season_folder, year_suffix).
    season:   e.g. 'SPWS-2024-2025'
    weapon:   e.g. 'EPEE'
    gender:   e.g. 'M'
    age_cat:  e.g. 'V2'
    """
    # 'SPWS-2024-2025' → '2024_25'
    parts = season.removeprefix("SPWS-").split("-")
    season_folder = f"{parts[0]}_{parts[1][2:]}"
    # 'SPWS-2024-2025' → '2024-2025'
    year_suffix = season.removeprefix("SPWS-")
    # 'V2', 'M', 'EPEE' → 'v2_m_epee'
    category_slug = f"{age_cat}_{gender}_{weapon}".lower()
    out_path = Path(f"supabase/data/{season_folder}/{category_slug}.sql")
    return out_path, season_folder, year_suffix


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generate per-category season seed SQL from an SPWS Excel workbook."
    )
    parser.add_argument("--xlsx",    required=True, help="Path to the Excel workbook")
    parser.add_argument("--season",  required=True, help="Season code, e.g. SPWS-2024-2025")
    parser.add_argument("--weapon",  required=True, choices=["EPEE", "FOIL", "SABRE"])
    parser.add_argument("--gender",  required=True, choices=["M", "F"])
    parser.add_argument("--age-cat", required=True, choices=["V0", "V1", "V2", "V3", "V4"],
                        dest="age_cat")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    season    = args.season
    weapon    = args.weapon
    gender    = args.gender
    age_cat   = args.age_cat

    out_path, season_folder, year_suffix = derive_paths(season, weapon, gender, age_cat)
    category_slug = f"{age_cat}_{gender}_{weapon}".lower()

    print(f"Reading {xlsx_path} ...")
    wb_data  = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_links = openpyxl.load_workbook(xlsx_path)  # for hyperlinks

    print(f"Connecting to DB ...")
    conn = psycopg2.connect(DB_URL)
    fencers = load_fencers(conn)
    print(f"  Loaded {len(fencers)} fencers from DB")
    conn.close()

    out_path.parent.mkdir(parents=True, exist_ok=True)

    xlsx_name = xlsx_path.name
    lines = []
    lines += [
        "-- =========================================================================",
        f"-- Season {year_suffix} — {age_cat} {gender} {weapon} — generated from {xlsx_name}",
        "-- Auto-loaded by supabase db reset via config.toml sql_paths glob.",
        "-- One file per age category per season; see supabase/data/{season}/{cat}.sql",
        "-- =========================================================================",
        "",
    ]

    total_matched   = 0
    total_unmatched = 0
    tournament_codes = []

    for sheet_name, (ttype, code, human_name) in SHEET_MAP.items():
        if sheet_name not in wb_data.sheetnames:
            print(f"  SKIP {sheet_name}: not found in workbook")
            continue

        data = extract_sheet(wb_data, wb_links, sheet_name)
        loc  = data["location"] or "?"
        dt   = data["date"]
        n    = data["n"]
        url  = data["url"]

        # Skip tournaments with N=0 (e.g. PEW10 Graz — no participants)
        if n == 0:
            print(f"  SKIP {sheet_name}: N=0 (no participants)")
            lines += [
                f"-- SKIP {sheet_name} ({human_name}): N=0 — tournament had no participants",
                "",
            ]
            continue

        event_code = f"{code}-{year_suffix}"
        tourn_code = f"{code}-{age_cat}-{gender}-{weapon}-{year_suffix}"
        tournament_codes.append(tourn_code)

        dt_sql   = sq(dt)
        url_sql  = sq(url)
        loc_raw  = data["location"]
        loc_sql  = sq(loc_raw) if loc_raw else "NULL"

        organizer_raw = "EVF" if ttype in ("PEW", "MEW") else "SPWS"
        loc_comment = f" ({loc})" if loc and loc != "?" else ""

        lines += [
            f"-- ---- {sheet_name}: {human_name}{loc_comment} ----",
            f"INSERT INTO tbl_event (txt_code, txt_name, txt_location, id_season, id_organizer, enum_status)",
            f"SELECT",
            f"    {sq(event_code)},",
            f"    {sq(human_name)},",
            f"    {loc_sql},",
            f"    (SELECT id_season FROM tbl_season WHERE txt_code = {sq(season)}),",
            f"    (SELECT id_organizer FROM tbl_organizer WHERE txt_code = {sq(organizer_raw)}),",
            f"    'COMPLETED'",
            f"WHERE NOT EXISTS (SELECT 1 FROM tbl_event WHERE txt_code = {sq(event_code)});",
            f"INSERT INTO tbl_tournament (",
            f"    id_event, txt_code, txt_name, enum_type,",
            f"    enum_weapon, enum_gender, enum_age_category,",
            f"    dt_tournament, int_participant_count, url_results,",
            f"    enum_import_status",
            f") VALUES (",
            f"    (SELECT id_event FROM tbl_event WHERE txt_code = {sq(event_code)}),",
            f"    {sq(tourn_code)},",
            f"    {sq(human_name)},",
            f"    '{ttype}',",
            f"    '{weapon}', '{gender}', '{age_cat}',",
            f"    {dt_sql}, {n}, {url_sql},",
            f"    'SCORED'",
            f");",
        ]

        matched_in_tournament = 0
        unmatched_in_tournament = 0

        for row in data["results"]:
            name  = row["name"]
            place = row["place"]
            match = fuzzy_match(name, fencers)

            if match:
                fid, matched_name, score = match
                matched_in_tournament += 1
                lines += [
                    f"INSERT INTO tbl_result (id_fencer, id_tournament, int_place, txt_scraped_name)",
                    f"VALUES (",
                    f"    {fid},",
                    f"    (SELECT id_tournament FROM tbl_tournament WHERE txt_code = {sq(tourn_code)}),",
                    f"    {place},",
                    f"    {sq(name)}",
                    f"); -- matched: {matched_name} (score={score})",
                ]
            else:
                unmatched_in_tournament += 1
                lines += [
                    f"-- UNMATCHED (score<{MATCH_THRESHOLD}): {sq(name)} place={place}",
                ]

        total_matched   += matched_in_tournament
        total_unmatched += unmatched_in_tournament
        print(f"  {sheet_name}: N={n}, results={len(data['results'])}, matched={matched_in_tournament}, unmatched={unmatched_in_tournament}")

        lines += [
            f"-- Compute scores for {tourn_code}",
            f"SELECT fn_calc_tournament_scores(",
            f"    (SELECT id_tournament FROM tbl_tournament WHERE txt_code = {sq(tourn_code)})",
            f");",
            "",
        ]

    lines += [
        "-- Summary",
        f"-- Total results matched:   {total_matched}",
        f"-- Total results unmatched: {total_unmatched}",
        "",
    ]

    sql = "\n".join(lines)
    out_path.write_text(sql, encoding="utf-8")
    print(f"\nWrote {out_path}")
    print(f"Total matched: {total_matched}, unmatched: {total_unmatched}")


if __name__ == "__main__":
    main()
