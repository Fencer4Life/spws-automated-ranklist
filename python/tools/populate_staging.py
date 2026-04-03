"""
Populate staging spreadsheet — data collection from source files.

Discovers Excel (2023-24, 2024-25) and XML (2025-26) source files,
extracts tournament metadata, fencer lists, and event information,
then returns data dicts ready for staging_spreadsheet.py tab builders.

Usage (called from staging_spreadsheet.py populate mode):
    from tools.populate_staging import collect_populate_data
    data = collect_populate_data(input_dir)
    # data = {"seasons": [...], "events": [...], "tournaments": [...], "fencers": [...]}
"""

from __future__ import annotations

import re
from pathlib import Path

# Weapon name mapping (Polish → English)
WEAPON_MAP = {"SZPADA": "EPEE", "FLORET": "FOIL", "SZABLA": "SABRE"}

# Season directories we process (keyed by season code)
SEASON_DIRS = {
    "SPWS-2023-2024": "Sezon 2023 - 2024",
    "SPWS-2024-2025": "Sezon 2024 - 2025",
    "SPWS-2025-2026": "Sezon_2025-2026",
}

XML_SEASON_DIR = "Sezon_2025-2026"
XML_SEASON_CODE = "SPWS-2025-2026"

# Filename pattern: {WEAPON}-{K?}{N}-{SEASON}.xlsx
_EXCEL_FILENAME_RE = re.compile(
    r"^(SZPADA|FLORET|SZABLA)-(K?)(\d)-(\d{4}-\d{4})\.xlsx$"
)


def parse_excel_filename(filename: str) -> dict | None:
    """Parse Excel filename into weapon/gender/age_cat/season_code.

    Args:
        filename: Just the filename (no directory), e.g. "SZPADA-2-2024-2025.xlsx"

    Returns:
        dict with weapon, gender, age_cat, season_code — or None if not a ranking file.
    """
    match = _EXCEL_FILENAME_RE.match(filename)
    if not match:
        return None

    weapon_pl, k_prefix, n, season = match.groups()
    return {
        "weapon": WEAPON_MAP[weapon_pl],
        "gender": "F" if k_prefix == "K" else "M",
        "age_cat": f"V{n}",
        "season_code": f"SPWS-{season}",
    }


def discover_excel_files(input_dir: Path) -> list[dict]:
    """Find all ranking Excel files for seasons 2023-24 and 2024-25.

    Returns:
        list[dict] with keys: path, weapon, gender, age_cat, season_code.
    """
    results = []
    for season_code, dir_name in SEASON_DIRS.items():
        season_dir = input_dir / dir_name
        if not season_dir.exists():
            continue
        for xlsx_path in sorted(season_dir.rglob("*.xlsx")):
            parsed = parse_excel_filename(xlsx_path.name)
            if parsed is None:
                continue
            parsed["path"] = xlsx_path
            results.append(parsed)
    return results


def discover_xml_files(input_dir: Path) -> list[dict]:
    """Find all result XML files for season 2025-26, skipping ELIMINACJE.

    Returns:
        list[dict] with keys: path, weapon, gender, categories, alt_name,
        title, date, season_code.
    """
    from scrapers.fencingtime_xml import parse_xml_metadata, detect_categories_from_altname

    xml_dir = input_dir / XML_SEASON_DIR
    if not xml_dir.exists():
        return []

    results = []
    for xml_path in sorted(xml_dir.rglob("*.xml")):
        file_bytes = xml_path.read_bytes()
        meta = parse_xml_metadata(file_bytes)

        # Skip ELIMINACJE (preliminary qualification rounds)
        if "ELIMINACJE" in meta.get("alt_name", "").upper():
            continue

        categories = detect_categories_from_altname(meta["alt_name"])
        results.append({
            "path": xml_path,
            "weapon": meta["weapon"],
            "gender": meta["gender"],
            "categories": categories,
            "alt_name": meta["alt_name"],
            "title": meta.get("title", ""),
            "date": meta.get("date", ""),
            "season_code": XML_SEASON_CODE,
        })

    return results


def extract_excel_tournaments(file_info: dict) -> list[dict]:
    """Extract tournament metadata from one Excel workbook.

    Uses SHEET_MAP to iterate sheets, detect_layout to pick the right
    extractor, and returns one tournament dict per sheet with results.

    Args:
        file_info: dict from discover_excel_files (path, weapon, gender, age_cat, season_code).

    Returns:
        list[dict] with keys: event_prefix, season_code, weapon, gender, age_cat,
        tournament_type, dt_tournament, participant_count, source_file,
        import_status, result_url, original_source, notes.
    """
    import openpyxl
    from tools.import_events_from_excel import (
        SHEET_MAP, SKIP_SHEETS, detect_layout,
        extract_sheet_2022, extract_sheet_2021, derive_event_url,
    )

    xlsx_path = file_info["path"]
    wb_data = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_links = openpyxl.load_workbook(xlsx_path)

    layout = detect_layout(wb_data)
    extract_fn = extract_sheet_2021 if layout == "2021" else extract_sheet_2022

    tournaments = []
    for sheet_name, (ttype, code_prefix, _human_name) in SHEET_MAP.items():
        if sheet_name not in wb_data.sheetnames:
            continue
        if sheet_name in SKIP_SHEETS:
            continue

        data = extract_fn(wb_data, wb_links, sheet_name)

        # Determine import status
        if data["result_count"] > 0:
            status = "SCORED"
        else:
            status = "EMPTY"

        tournaments.append({
            "event_prefix": code_prefix,
            "season_code": file_info["season_code"],
            "weapon": file_info["weapon"],
            "gender": file_info["gender"],
            "age_cat": file_info["age_cat"],
            "type": ttype,
            "dt_tournament": data["date_start"],
            "participant_count": data["participant_count"],
            "source_file": str(xlsx_path),
            "import_status": status,
            "result_url": derive_event_url(data["url"]),
            "original_source": data["url"],
            "notes": None,
        })

    wb_data.close()
    wb_links.close()
    return tournaments


def extract_xml_tournaments(file_info: dict) -> list[dict]:
    """Extract tournament metadata from one XML file.

    Creates one tournament row per category. For combined categories
    (e.g., V0+V1), the total participant count is shared and notes
    indicate the combination.

    Args:
        file_info: dict from discover_xml_files (path, weapon, gender,
                   categories, alt_name, title, date, season_code).

    Returns:
        list[dict] with same keys as extract_excel_tournaments.
    """
    from scrapers.fencingtime_xml import parse_fencingtime_xml
    from tools.import_events_from_excel import parse_date

    xml_path = file_info["path"]
    file_bytes = xml_path.read_bytes()
    results = parse_fencingtime_xml(file_bytes)
    participant_count = len(results)
    status = "SCORED" if participant_count > 0 else "EMPTY"

    # Derive event prefix from TitreLong (Roman numeral → PPn)
    event_prefix = _event_prefix_from_title(file_info["title"])

    dt_tournament = parse_date(file_info["date"])

    categories = file_info["categories"]
    is_combined = len(categories) > 1
    combined_note = f"combined {''.join(categories)}" if is_combined else None

    tournaments = []
    for cat in categories:
        tournaments.append({
            "event_prefix": event_prefix,
            "season_code": file_info["season_code"],
            "weapon": file_info["weapon"],
            "gender": file_info["gender"],
            "age_cat": cat,
            "type": "PPW",
            "dt_tournament": dt_tournament,
            "participant_count": participant_count,
            "source_file": str(xml_path),
            "import_status": status,
            "result_url": None,
            "original_source": None,
            "notes": combined_note,
        })

    return tournaments


# Roman numeral → integer for event prefix parsing
_ROMAN = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7, "VIII": 8}


def _event_prefix_from_title(title: str) -> str:
    """Extract event prefix from TitreLong.

    "IV Puchar Polski Weteranów Gdańsk 2026" → "PPW4"
    """
    match = re.match(r"(I{1,3}V?|VI{0,3}|VIII?)\s+", title)
    if match:
        roman = match.group(1)
        n = _ROMAN.get(roman)
        if n is not None:
            return f"PPW{n}"
    return "PPW0"


def build_events_from_tournaments(tournaments: list[dict]) -> list[dict]:
    """Group tournaments by (event_prefix, season_code) → deduplicated event rows.

    Merges date ranges and fills missing URLs across tournament rows
    that belong to the same event.

    Returns:
        list[dict] with keys: event_prefix, season_code, name,
        dt_start, dt_end, organizer, status.
    """
    grouped: dict[tuple[str, str], dict] = {}
    for t in tournaments:
        key = (t["event_prefix"], t["season_code"])
        if key not in grouped:
            grouped[key] = {
                "event_prefix": t["event_prefix"],
                "season_code": t["season_code"],
                "name": t["event_prefix"],
                "dt_start": t["dt_tournament"],
                "dt_end": t["dt_tournament"],
                "organizer": "SPWS" if t["type"] in ("PPW", "MPW") else "EVF",
                "status": "COMPLETED",
            }
        else:
            evt = grouped[key]
            dt = t["dt_tournament"]
            if dt:
                if evt["dt_start"] is None or dt < evt["dt_start"]:
                    evt["dt_start"] = dt
                if evt["dt_end"] is None or dt > evt["dt_end"]:
                    evt["dt_end"] = dt

    return list(grouped.values())


def determine_import_status(result_count: int | None) -> str:
    """Determine import status from result count.

    Args:
        result_count: Number of result rows, or None if sheet is missing.

    Returns:
        "SCORED" if results exist, "EMPTY" if zero results, "LOST" if None.
    """
    if result_count is None:
        return "LOST"
    return "SCORED" if result_count > 0 else "EMPTY"


def extract_fencers_from_results(tournaments: list[dict]) -> list[dict]:
    """Collect unique fencers from SPWS-organized tournament source files.

    Only includes fencers who participated in PPW, GPW, or MPW tournaments
    (SPWS-organized events). Fencers who only appear in international events
    (PEW, MEW, PSW, MSW) are excluded.

    Args:
        tournaments: list of tournament dicts with source_file and import_status.

    Returns:
        list[dict] with keys: surname, first_name, birth_year (int or None),
        source_note (str).
    """
    from scrapers.fencingtime_xml import parse_fencingtime_xml_enriched

    # Only extract fencers from SPWS-organized tournament types
    _SPWS_TYPES = {"PPW", "GPW", "MPW"}

    seen: dict[tuple[str, str], dict] = {}  # (surname_upper, first_upper) → fencer

    for t in tournaments:
        if t.get("type") not in _SPWS_TYPES:
            continue
        if t["import_status"] != "SCORED":
            continue
        source = t.get("source_file", "")
        if not source:
            continue

        source_path = Path(source)
        if not source_path.exists():
            continue

        if source_path.suffix.lower() == ".xml":
            file_bytes = source_path.read_bytes()
            results = parse_fencingtime_xml_enriched(file_bytes)
            for r in results:
                parts = r["fencer_name"].split(None, 1)
                surname = parts[0] if parts else ""
                first_name = parts[1] if len(parts) > 1 else ""
                birth_year = None
                if r.get("birth_date"):
                    birth_year = int(r["birth_date"][:4])
                key = (surname.upper(), first_name.upper())
                if key not in seen:
                    seen[key] = {
                        "surname": surname,
                        "first_name": first_name,
                        "birth_year": birth_year,
                        "source_note": source_path.name,
                    }
                elif birth_year and not seen[key].get("birth_year"):
                    seen[key]["birth_year"] = birth_year

        elif source_path.suffix.lower() == ".xlsx":
            _extract_fencers_from_excel(source_path, seen)

    return list(seen.values())


# Cache to avoid reopening the same workbook for multiple tournaments
_excel_fencer_cache: set[str] = set()


def _extract_fencers_from_excel(xlsx_path: Path, seen: dict) -> None:
    """Extract fencer names from PPW/GPW/MPW sheets in an Excel workbook."""
    import openpyxl
    from tools.import_events_from_excel import (
        SHEET_MAP, SKIP_SHEETS, detect_layout,
    )

    _SPWS_TYPES = {"PPW", "GPW", "MPW"}

    # Skip if already processed this file
    path_str = str(xlsx_path)
    if path_str in _excel_fencer_cache:
        return
    _excel_fencer_cache.add(path_str)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    layout = detect_layout(wb)
    name_col = 3  # column C has fencer names in both layouts
    start_row = 5 if layout == "2021" else 6

    for sheet_name, (ttype, _code, _name) in SHEET_MAP.items():
        if ttype not in _SPWS_TYPES:
            continue
        if sheet_name not in wb.sheetnames or sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        for r in range(start_row, ws.max_row + 1):
            name_val = ws.cell(r, name_col).value
            if not name_val or str(name_val).strip() in ("x", "X", ""):
                continue
            name_str = str(name_val).strip()
            parts = name_str.split(None, 1)
            surname = parts[0] if parts else ""
            first_name = parts[1] if len(parts) > 1 else ""
            key = (surname.upper(), first_name.upper())
            if key not in seen:
                seen[key] = {
                    "surname": surname,
                    "first_name": first_name,
                    "birth_year": None,
                    "source_note": xlsx_path.name,
                }


MEMBERS_FILENAME = "czlonkowie SPWS.xlsx"
MEMBERS_SHEET = "Kategorie wiekowe"


def _enrich_birth_years_from_members(fencers: list[dict], input_dir: Path) -> int:
    """Enrich fencer birth years from the SPWS members spreadsheet.

    Reads the 'Kategorie wiekowe' sheet from czlonkowie SPWS.xlsx and
    fills in birth_year for fencers that match by name but lack a birth year.

    Returns:
        Number of fencers enriched.
    """
    import openpyxl

    members_path = input_dir / MEMBERS_FILENAME
    if not members_path.exists():
        return 0

    wb = openpyxl.load_workbook(members_path, data_only=True)
    ws = wb[MEMBERS_SHEET]

    # Build lookup: (SURNAME_UPPER, FIRST_UPPER) → birth_year
    member_years: dict[tuple[str, str], int] = {}
    for r in range(2, ws.max_row + 1):
        name_val = ws.cell(r, 2).value
        if not name_val:
            continue
        name_str = str(name_val).strip()
        parts = name_str.split(None, 1)
        surname = parts[0] if parts else ""
        first_name = parts[1] if len(parts) > 1 else ""
        birth_year = ws.cell(r, 4).value
        if birth_year:
            member_years[(surname.upper(), first_name.upper())] = int(birth_year)

    wb.close()

    enriched = 0
    for f in fencers:
        if f.get("birth_year"):
            continue
        key = (f["surname"].upper(), f["first_name"].upper())
        year = member_years.get(key)
        if year:
            f["birth_year"] = year
            f["source_note"] = f.get("source_note", "") or MEMBERS_FILENAME
            enriched += 1

    return enriched


def collect_populate_data(input_dir: Path) -> dict:
    """Orchestrator: discover files, extract tournaments/fencers, build events.

    Returns:
        dict with keys: seasons, events, tournaments, fencers.
        Each value is a list of dicts matching the tab builder format.
    """
    # Seasons (static for now — 3 seasons)
    seasons = [
        {
            "season_code": "SPWS-2023-2024",
            "dt_start": "2023-09-01", "dt_end": "2024-08-31",
            "bool_active": False,
            "ppw_best_count": 4, "ppw_total_rounds": 8,
            "mpw_multiplier": 1.5,
            "pew_best_count": 4, "mew_multiplier": 2.0,
            "msw_multiplier": 2.0, "import_log": "",
        },
        {
            "season_code": "SPWS-2024-2025",
            "dt_start": "2024-09-01", "dt_end": "2025-08-31",
            "bool_active": False,
            "ppw_best_count": 4, "ppw_total_rounds": 5,
            "mpw_multiplier": 1.5,
            "pew_best_count": 4, "mew_multiplier": 2.0,
            "msw_multiplier": 2.0, "import_log": "",
        },
        {
            "season_code": "SPWS-2025-2026",
            "dt_start": "2025-09-01", "dt_end": "2026-08-31",
            "bool_active": True,
            "ppw_best_count": 4, "ppw_total_rounds": 5,
            "mpw_multiplier": 1.5,
            "pew_best_count": 4, "mew_multiplier": 2.0,
            "msw_multiplier": 2.0, "import_log": "",
        },
    ]

    # Discover files
    excel_files = discover_excel_files(input_dir)
    xml_files = discover_xml_files(input_dir)

    # Extract tournaments from all files
    all_tournaments = []
    for f in excel_files:
        all_tournaments.extend(extract_excel_tournaments(f))
    for f in xml_files:
        all_tournaments.extend(extract_xml_tournaments(f))

    # Build deduplicated events
    events = build_events_from_tournaments(all_tournaments)

    # Extract fencers from all tournament sources
    fencers = extract_fencers_from_results(all_tournaments)

    # Enrich birth years from SPWS members spreadsheet
    _enrich_birth_years_from_members(fencers, input_dir)

    return {
        "seasons": seasons,
        "events": events,
        "tournaments": all_tournaments,
        "fencers": fencers,
    }
