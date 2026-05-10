"""
Recreate SPWS active season 2025-2026 from primary source artifacts.

PPW1: 30 XLSX files in doc/external_files/Sezon_2025-2026/XLS/
PPW2: FTL event URL (url_event)
PPW3: FTL event URL (url_event) — full event rescrape
PPW4: 17 XML files in doc/external_files/Sezon_2025-2026/PPW4-GDANSK/
PPW5: 21 XML files in doc/external_files/Sezon_2025-2026/PPW5-GDANSK/
MPW: skipped (no source — championship hasn't run yet)

Bulk run, halts on first failure (pipeline halt, exception, or commit RPC error).

Usage:
  python -m python.tools.recreate_active_season_2025_2026                 # all events
  python -m python.tools.recreate_active_season_2025_2026 --event PPW1    # just PPW1
  python -m python.tools.recreate_active_season_2025_2026 --no-commit     # stage only
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import io

from python.pipeline.db_connector import create_db_connector
from python.pipeline.draft_store import DraftStore
from python.pipeline.review_cli import Fetcher, ReviewSession, _annotate_parsed
from python.tools.phase5_runner import (
    _fetch_event_meta,
    _coerce_date,
    _stamp_event_metadata,
    _consolidate_duplicate_codes,
    _drop_empty_tournament_drafts,
    _check_pool_round_count,
    _multi_summary_md,
)


# ---------------------------------------------------------------------------
# Source paths (relative to repo root)
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parents[2]
XLS_DIR = ROOT / "doc" / "external_files" / "Sezon_2025-2026" / "XLS"
PPW4_XML_DIR = ROOT / "doc" / "external_files" / "Sezon_2025-2026" / "PPW4-GDANSK"
PPW5_XML_DIR = ROOT / "doc" / "external_files" / "Sezon_2025-2026" / "PPW5-GDANSK"


# ---------------------------------------------------------------------------
# PPW1 XLSX: filename → (weapon, gender, age_category)
# Examples:
#   FLORET-0-2025-2026.xlsx  → FOIL, M, V0
#   FLORET-K2-2025-2026.xlsx → FOIL, F, V2
#   SZABLA-3-2025-2026.xlsx  → SABRE, M, V3
#   SZPADA-K4-2025-2026.xlsx → EPEE, F, V4
# ---------------------------------------------------------------------------
WEAPON_PL_TO_EN = {"FLORET": "FOIL", "SZABLA": "SABRE", "SZPADA": "EPEE"}


def _xlsx_metadata_from_filename(filename: str) -> tuple[str, str, str]:
    stem = Path(filename).stem
    parts = stem.split("-")
    weapon_pl = parts[0]
    cat_part = parts[1]
    weapon = WEAPON_PL_TO_EN.get(weapon_pl)
    if weapon is None:
        raise ValueError(f"unknown weapon prefix in {filename!r}")
    if cat_part.startswith("K"):
        gender = "F"
        cat_n = cat_part[1:]
    else:
        gender = "M"
        cat_n = cat_part
    if cat_n not in {"0", "1", "2", "3", "4"}:
        raise ValueError(f"unknown V-cat in {filename!r}: {cat_part}")
    return weapon, gender, f"V{cat_n}"


def _ppw1_xlsx_files() -> list[Path]:
    skip_tokens = ("SuperFive", "template", "czlonkowie", "Template")
    files = sorted(XLS_DIR.glob("*.xlsx"))
    files = [f for f in files if not any(s in f.stem for s in skip_tokens)]
    return files


# ---------------------------------------------------------------------------
# PPW1 XLSX → ParsedTournament
#
# These workbooks have multiple sheets (Ranking + PP1..PP5 + MPW). We read the
# `PP1` sheet only — that's the source-of-truth result table for the PPW1
# tournament corresponding to this weapon/gender/V-cat. Headers contain
# multi-line strings ("Miejsce\nPlace", "Nazwisko Imię\nName", etc.) so we
# match by substring rather than exact token.
# ---------------------------------------------------------------------------

def _parse_ppw1_xlsx(file_path: Path):
    """Read sheet `PP1` of a SPWS PPW workbook → ParsedTournament IR."""
    import openpyxl
    from python.pipeline.ir import (
        ParsedResult, ParsedTournament, SourceKind, make_synthetic_id,
    )

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if "PP1" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"{file_path.name}: no PP1 sheet (sheets: {wb.sheetnames})")
    ws = wb["PP1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header row by substring match on "Miejsce" (place) + "Nazwisko" (name).
    header_idx = None
    col_place = col_name = col_country = None
    for i, row in enumerate(rows):
        norm = []
        for c in row:
            if c is None:
                norm.append("")
            else:
                norm.append(str(c).strip().lower())
        # Find first cell that contains "miejsce" (Polish for place)
        place_idx = next((j for j, c in enumerate(norm) if "miejsce" in c), None)
        name_idx = next((j for j, c in enumerate(norm) if "nazwisko" in c), None)
        if place_idx is not None and name_idx is not None:
            header_idx = i
            col_place = place_idx
            col_name = name_idx
            country_idx = next(
                (j for j, c in enumerate(norm) if "państwo" in c or "country" in c),
                None,
            )
            col_country = country_idx
            break

    if header_idx is None:
        raise ValueError(f"{file_path.name}: PP1 sheet has no recognizable header")

    # ADR-066 walkover: pre-scan for non-empty fencer-name rows.
    # If exactly one such row has no place, treat it as place=1.
    candidates: list[tuple[int, tuple, str]] = []  # (row_index, row, name)
    for i, row in enumerate(rows[header_idx + 1:], start=1):
        if col_place >= len(row) or col_name >= len(row):
            continue
        name_val = row[col_name]
        if name_val is None:
            continue
        name = str(name_val).strip()
        if not name:
            continue
        candidates.append((i, row, name))
    is_walkover = len(candidates) == 1

    parsed_results = []
    for i, row, name in candidates:
        place_val = row[col_place]
        if place_val is None:
            if is_walkover:
                place = 1
            else:
                continue
        else:
            try:
                place = int(float(str(place_val)))
            except (ValueError, TypeError):
                continue
        country = None
        if col_country is not None and col_country < len(row) and row[col_country]:
            country = str(row[col_country]).strip() or None
        parsed_results.append(ParsedResult(
            source_row_id=make_synthetic_id(
                SourceKind.FILE_IMPORT, row_index=i, place=place, name=name,
            ),
            fencer_name=name,
            place=place,
            fencer_country=country,
        ))

    return ParsedTournament(
        source_kind=SourceKind.FILE_IMPORT,
        results=parsed_results,
        raw_pool_size=len(parsed_results),
        source_url=None,
        source_artifact_path=str(file_path),
    )


# ---------------------------------------------------------------------------
# Per-event ingestion driver
# ---------------------------------------------------------------------------

def _iter_ppw1(fetcher: Fetcher):
    files = _ppw1_xlsx_files()
    if len(files) != 30:
        raise ValueError(
            f"PPW1 expected 30 ranking XLSX files, found {len(files)} in {XLS_DIR}"
        )
    for f in files:
        weapon, gender, age = _xlsx_metadata_from_filename(f.name)
        parsed = _parse_ppw1_xlsx(f)
        annotated = _annotate_parsed(
            parsed, weapon=weapon, gender=gender,
            age_category=age, ftl_source_name=f.name,
        )
        yield annotated, f.name


import re as _re_xml
import dataclasses as _dc_xml

_COMPOUND_VCAT_RE = _re_xml.compile(r"^V\d(V\d)+$", _re_xml.IGNORECASE)


def _normalize_compound_vcat(parsed):
    """Set category_hint=None when it's a compound like 'V0V1', 'V0V1V2'.

    Reason: ADR-056 revision treats single-V-cat hint as authoritative
    placement; compound hints aren't real V-cats — they signal that the
    XML is a joint pool. Setting to None routes through BY-derivation.
    """
    hint = getattr(parsed, "category_hint", None)
    if hint and _COMPOUND_VCAT_RE.match(hint):
        try:
            return _dc_xml.replace(parsed, category_hint=None)
        except (TypeError, ValueError):
            setattr(parsed, "category_hint", None)
    return parsed


def _split_mixed_gender_xml(file_path: Path):
    """Yield 1 or 2 ParsedTournament IRs from an XML file.

    For single-gender brackets (root Sexe in {'M','F'}): yield 1 IR via
    the standard `fencingtime_xml.parse` path.

    For mixed-gender brackets (root Sexe='X', e.g. RESULTS_GRVETXE...):
    yield 2 IRs — one per gender — split by each `Tireur`'s own `Sexe`
    attribute. Places are re-ranked within each gender (1..N_gender).
    Pipeline can then apply ADR-056 BY-derivation per fencer for V-cat.
    """
    import xml.etree.ElementTree as ET
    from python.scrapers.fencingtime_xml import parse as ft_xml_parse

    data = file_path.read_bytes()
    root = ET.fromstring(data)
    bracket_sexe = root.attrib.get("Sexe", "")

    parsed = ft_xml_parse(data, source_url=f"file://{file_path}")

    if bracket_sexe in ("M", "F"):
        # Standard single-gender — pass through.
        yield parsed
        return

    # Mixed bracket — build name→gender map from per-Tireur Sexe attrs.
    name_to_gender: dict[str, str] = {}
    for t in root.findall(".//Tireur"):
        nom = t.attrib.get("Nom", "")
        prenom = t.attrib.get("Prenom", "")
        sexe = t.attrib.get("Sexe", "")
        if sexe in ("M", "F"):
            full = f"{nom} {prenom}".strip() if prenom else nom
            name_to_gender[full] = sexe

    by_gender: dict[str, list] = {"M": [], "F": []}
    for r in parsed.results:
        g = name_to_gender.get(r.fencer_name)
        if g in ("M", "F"):
            by_gender[g].append(r)

    for g in ("M", "F"):
        rs = by_gender[g]
        if not rs:
            continue
        # Re-rank within gender (1..N) preserving original-place ordering.
        rs_sorted = sorted(rs, key=lambda r: r.place)
        renumbered = [_dc_xml.replace(r, place=i + 1) for i, r in enumerate(rs_sorted)]
        new_parsed = _dc_xml.replace(
            parsed,
            results=renumbered,
            gender=g,
            raw_pool_size=len(renumbered),
        )
        yield new_parsed


def _iter_xml_dir(fetcher: Fetcher, xml_dir: Path):
    files = sorted(xml_dir.glob("*.xml"))
    if not files:
        raise ValueError(f"no XML files in {xml_dir}")
    for f in files:
        for parsed in _split_mixed_gender_xml(f):
            parsed = _normalize_compound_vcat(parsed)
            label = f.name if getattr(parsed, "gender", None) in (None,) else \
                    f"{f.name} (split→{parsed.gender})"
            yield parsed, label


def _iter_url_event(fetcher: Fetcher, event_url: str):
    parsed_list, splitter_skips = fetcher.fetch_event_url_with_skips(event_url)
    for p in parsed_list:
        yield p, getattr(p, "_ftl_source_name", None) or "?"
    # splitter_skips returned out-of-band — caller handles via session/pool tracking


def _dedupe_result_drafts(db, run_id: str) -> int:
    """Drop duplicate (id_fencer, id_tournament_draft) result_draft rows.

    Keeps the row with the lowest int_place per group (best result wins).
    Returns the count of deleted rows.

    Why this is needed: when ADR-065's V-cat marker check downgrades a
    misregistered bracket to joint-pool, BY-derivation routes fencers to
    their actual V-cats — which can collide with per-V-cat-bracket drafts
    already present in the run. _consolidate_duplicate_codes merges
    tournament_draft rows by txt_code but does not dedupe result_draft
    rows that share a fencer within a merged tournament. Without this
    step, fn_commit_event_draft fails with uq_result_fencer_tournament.
    """
    import subprocess
    sql = f"""
WITH ranked AS (
  SELECT id_result_draft,
         id_fencer, id_tournament_draft, int_place,
         ROW_NUMBER() OVER (
           PARTITION BY id_fencer, id_tournament_draft
           ORDER BY int_place NULLS LAST, id_result_draft
         ) AS rn
  FROM tbl_result_draft
  WHERE txt_run_id = '{run_id}'
),
del AS (
  DELETE FROM tbl_result_draft
  WHERE id_result_draft IN (SELECT id_result_draft FROM ranked WHERE rn > 1)
  RETURNING 1
)
SELECT COUNT(*) FROM del;
"""
    r = subprocess.run(
        ['docker', 'exec', '-i', 'supabase_db_SPWSranklist',
         'psql', '-U', 'postgres', '-tA', '-c', sql],
        capture_output=True, text=True, timeout=60,
    )
    if r.returncode != 0:
        print(f"  ⚠ dedupe failed: {r.stderr[-200:]}", flush=True)
        return 0
    n = int(r.stdout.strip().split("\n")[-1] or "0")
    if n:
        print(f"  ↻ deduped {n} duplicate (id_fencer, id_tournament_draft) "
              f"result_drafts (kept best place per fencer)", flush=True)
    return n


EVENT_HANDLERS = {
    "PPW1-2025-2026": ("xlsx",  None),
    "PPW2-2025-2026": ("url",   None),
    "PPW3-2025-2026": ("url",   None),
    "PPW4-2025-2026": ("xml",   PPW4_XML_DIR),
    "PPW5-2025-2026": ("xml",   PPW5_XML_DIR),
}

CHRONO_ORDER = [
    "PPW1-2025-2026",
    "PPW2-2025-2026",
    "PPW3-2025-2026",
    "PPW4-2025-2026",
    "PPW5-2025-2026",
]


def ingest_event(event_code: str, *, db, do_commit: bool) -> tuple[str | None, bool, str]:
    """
    Stage drafts for one event from its primary source. Optionally commit.

    Returns (run_id, ok, reason). ok=False with reason='halt' / 'error' / 'commit_failed'.
    """
    handler_kind, xml_dir = EVENT_HANDLERS[event_code]
    fetcher = Fetcher()
    store = DraftStore(db)
    session = ReviewSession(
        event_code=event_code, db=db, draft_store=store,
        season_end_year=2026, fetcher=fetcher,
    )

    event_meta = _fetch_event_meta(db, event_code)
    parsed_date = _coerce_date(event_meta["dt_start"])
    city_default = event_meta.get("txt_location")
    country_default = "PL" if event_meta["organizer_code"] == "SPWS" else None

    # ADR-066: resolve the per-season minimum-participants threshold once
    # per event; gate every parsed bracket against it inside the loop.
    from python.pipeline.db_connector import (
        derive_tourn_type_from_event_code,
        gate_below_min_participants,
    )
    id_season = event_meta["_full_row"]["id_season"]
    tourn_type = derive_tourn_type_from_event_code(event_code)

    # Build the parsed-IR iterator
    if handler_kind == "xlsx":
        src_iter = _iter_ppw1(fetcher)
    elif handler_kind == "xml":
        src_iter = _iter_xml_dir(fetcher, xml_dir)
    elif handler_kind == "url":
        if not event_meta["urls"]:
            return None, False, f"{event_code}: no url_event* set"
        slot, url = event_meta["urls"][0]
        src_iter = _iter_url_event(fetcher, url)
    else:
        return None, False, f"unknown handler {handler_kind!r}"

    ctxs: list = []
    pool_brackets: list[dict] = []
    halts = 0
    errs = 0
    drafted = 0

    print(f"\n=== {event_code} ({handler_kind}) ===", flush=True)
    print(f"  organizer={event_meta['organizer_code']} dt_start={event_meta['dt_start']} "
          f"city={city_default!r}", flush=True)

    for i, (parsed, label) in enumerate(src_iter, start=1):
        parsed = _stamp_event_metadata(
            parsed, parsed_date=parsed_date,
            country_default=country_default,
            city_default=city_default,
        )
        n_results = len(getattr(parsed, "results", []) or [])
        skip, reason = gate_below_min_participants(
            db, id_season, tourn_type, n_results,
        )
        if skip:
            print(f"  [{i:3d}] {label} — {reason}, skip", flush=True)
            ctxs.append((1, parsed, None, reason))
            continue
        print(f"  [{i:3d}] {label} — {getattr(parsed, 'weapon', '?')}/"
              f"{getattr(parsed, 'gender', '?')}/"
              f"{getattr(parsed, 'category_hint', getattr(parsed, 'age_category', '?'))} "
              f"n={n_results}", flush=True, end=" ")
        try:
            ctx, _ = session.run_iteration(parsed, staging_dir=Path("doc/staging"))
            ctxs.append((1, parsed, ctx, None))
            if ctx.halted:
                halts += 1
                print(f"HALT @ {ctx.halted_at_stage}: {ctx.halt_detail}", flush=True)
                if getattr(ctx, "is_pool_round", False):
                    pool_brackets.append({
                        "weapon": parsed.weapon or "?",
                        "name": label,
                        "url": getattr(parsed, "source_url", None),
                        "reason": getattr(ctx, "halt_detail", "") or "",
                    })
            else:
                drafted += 1
                print("OK", flush=True)
        except Exception as e:
            errs += 1
            print(f"EXC: {e}", flush=True)
            ctxs.append((1, parsed, None, str(e)))

    # Post-bracket consolidation (mirrors phase5_runner)
    if session.run_id:
        _consolidate_duplicate_codes(db, session.run_id)
        _drop_empty_tournament_drafts(db, session.run_id)
        # ADR-065 follow-on: after consolidation, drop duplicate
        # (id_fencer, id_tournament_draft) pairs from tbl_result_draft.
        # When the V-cat marker check downgrades a misregistered bracket
        # to joint-pool, BY-derivation can route fencers to V-cats that
        # already have per-V-cat-bracket drafts. Consolidation merges the
        # tournament_draft rows but keeps both result_draft rows for the
        # same fencer — uq_result_fencer_tournament rejects this on commit.
        # We dedupe by keeping the lowest int_place per (id_fencer, id_tournament_draft).
        _dedupe_result_drafts(db, session.run_id)

    pool_warnings = _check_pool_round_count(pool_brackets)
    for w in pool_warnings:
        print(f"  ⚠ {w}", flush=True)

    # Write per-event staging summary (.md)
    md = _multi_summary_md(
        event_code, event_meta, ctxs, db=db,
        pool_brackets=pool_brackets, pool_warnings=pool_warnings,
        run_id=session.run_id,
        url_check_results=getattr(session, "url_check_results", {}),
    )
    out_path = Path("doc/staging") / f"{event_code}.md"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(md)
    print(f"  → wrote {out_path}", flush=True)

    print(f"  summary: drafted={drafted} halted={halts} errors={errs} run_id={session.run_id}",
          flush=True)

    if not do_commit:
        return session.run_id, True, "stage-only"

    if session.run_id is None:
        return None, True, "no run_id (nothing drafted)"

    # ---------------------------------------------------------------------
    # Alias review gate — mirrors phase5_runner.
    # ---------------------------------------------------------------------
    # Phase 5 Option-1 stage-time flush (ADR-058+059): write EVERY pending
    # pair (✓, ❓, AND ❌) to tbl_fencer.json_name_aliases so the
    # FencerAliasManager UI surfaces them all for operator review. Wrong
    # matches reach the alias column briefly; the operator resolves via
    # Transfer / Discard / Create-new-fencer in the UI. This step is
    # what makes Phase-5 sign-off a real review gate — without it, the
    # UI shows nothing to act on and ❌ pairs are invisible.
    from python.pipeline.alias_writeback import (
        derive_pending_from_run_id, has_blocking_pairs,
        flush_pending_aliases,
    )
    stage_pending = derive_pending_from_run_id(db, session.run_id)
    if stage_pending:
        try:
            stage_flush = flush_pending_aliases(
                db, stage_pending, include_all=True,
            )
            print(
                f"  stage-time alias flush: "
                f"{stage_flush['written']} written to tbl_fencer "
                f"(all verdicts; UI will surface for review), "
                f"{len(stage_flush['errors'])} errors",
                flush=True,
            )
            if stage_flush["errors"]:
                for fid, alias, msg in stage_flush["errors"][:5]:
                    print(f"     ⚠ #{fid} {alias!r}: {msg}", flush=True)
        except Exception as e:
            print(f"  ⚠ stage-time alias flush failed: {e}", flush=True)

    # Re-derive pending AFTER the stage-time flush so user-confirmed
    # aliases (json_user_confirmed_aliases) are respected. Same call as
    # phase5_runner --commit-run-id uses.
    pending = derive_pending_from_run_id(db, session.run_id)

    # Pipeline halts block commit (e.g. pool-round-detected halts). Halt
    # check goes here — AFTER stage-flush so the UI is loaded with the
    # successful brackets' aliases for operator review.
    if halts > 0 or errs > 0:
        return session.run_id, False, (
            f"pipeline_halts={halts} errors={errs}; alias data flushed "
            f"to tbl_fencer for UI review at "
            f"http://localhost:5173/?admin=1 — resolve and re-stage"
        )
    blockers = [p for p in pending if p.icon == "❌"]
    ambiguous = [p for p in pending if p.icon == "❓"]
    confirmed = [p for p in pending if p.icon == "✓"]

    if blockers or ambiguous:
        # Halt — surface alias proposals for operator review in UI.
        import sys as _sys
        print(f"  ⛔ alias review required:", file=_sys.stderr)
        print(f"     ❌ {len(blockers)} blocking (probably wrong matches)", file=_sys.stderr)
        print(f"     ❓ {len(ambiguous)} ambiguous (uncertain)", file=_sys.stderr)
        print(f"     ✓ {len(confirmed)} would auto-write on resolution", file=_sys.stderr)
        for p in blockers[:20]:
            print(f"     ❌ #{p.id_fencer} {p.scraped_name!r} → {p.canonical!r} ({p.reason})",
                  file=_sys.stderr)
        for p in ambiguous[:20]:
            print(f"     ❓ #{p.id_fencer} {p.scraped_name!r} → {p.canonical!r} ({p.reason})",
                  file=_sys.stderr)
        print(
            f"  → resolve at http://localhost:5173/?admin=1 (FencerAliasManager UI), "
            f"then re-stage this event.",
            file=_sys.stderr,
        )
        return session.run_id, False, (
            f"alias_review_required (❌={len(blockers)}, ❓={len(ambiguous)})"
        )

    if confirmed:
        print(f"  flushing {len(confirmed)} ✓ aliases to tbl_fencer...", flush=True)
        flush_result = flush_pending_aliases(db, pending)
        print(f"  ✓ aliases written: {flush_result['written']}, "
              f"errors: {len(flush_result['errors'])}", flush=True)
        if flush_result["errors"]:
            for fid, alias, msg in flush_result["errors"][:5]:
                print(f"     ⚠ #{fid} {alias!r}: {msg}", flush=True)

    # Commit drafts to live tables
    print(f"  committing run {session.run_id}...", flush=True)
    try:
        resp = db._sb.rpc("fn_commit_event_draft", {"p_run_id": session.run_id}).execute()
        print(f"  ✓ committed: {resp.data}", flush=True)
        return session.run_id, True, "committed"
    except Exception as e:
        print(f"  ✗ commit failed: {e}", flush=True)
        return session.run_id, False, f"commit_failed: {e}"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--event", default="all",
                        help="PPW1 / PPW2 / PPW3 / PPW4 / PPW5 / all (default)")
    parser.add_argument("--no-commit", action="store_true",
                        help="Stage drafts only, do not commit")
    args = parser.parse_args()

    if args.event == "all":
        events = list(CHRONO_ORDER)
    else:
        target = f"{args.event}-2025-2026"
        if target not in EVENT_HANDLERS:
            print(f"unknown event: {args.event}", file=sys.stderr)
            return 1
        events = [target]

    db = create_db_connector()

    for ec in events:
        run_id, ok, reason = ingest_event(ec, db=db, do_commit=not args.no_commit)
        if not ok:
            print(f"\n⛔ HALT after {ec}: {reason}  (run_id={run_id})", file=sys.stderr)
            print(f"   → review doc/staging/{ec}.md, then "
                  f"`python -m python.tools.phase5_runner --event-code {ec} "
                  f"--commit-run-id {run_id}` after fixing.",
                  file=sys.stderr)
            return 2

    print("\n✓ All events ingested successfully", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
