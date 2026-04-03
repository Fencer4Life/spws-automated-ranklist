"""
import_events_from_excel.py
---------------------------
Scans all .xlsx files recursively in a folder tree and creates/updates events.
Assigns each event to the correct DB season by matching its date against
tbl_season.dt_start / dt_end ranges.

Usage:
    python python/tools/import_events_from_excel.py \
        --folder doc/external_files/ \
        --dry-run

    python python/tools/import_events_from_excel.py \
        --folder doc/external_files/

The script:
1. Loads all season date ranges from the DB
2. Recursively finds .xlsx files (skips SuperFive, czlonkowie, Backup)
3. Reads ONE representative file per season-folder (deduplicates across weapons)
4. Extracts event metadata from each sheet (GP1, PP1, PEW1, MPW, etc.)
5. Assigns each event to a DB season by its date
6. Generates event codes: {prefix}-{year_suffix} where year_suffix comes from the season
7. Upserts events into the DB (only fills NULL fields on existing events)
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import psycopg2

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DB_URL = "postgresql://postgres:postgres@127.0.0.1:54322/postgres"

# Map sheet name → (tournament_type, event_code_prefix, human_name)
# Covers all naming conventions across 2021-2025 seasons
SHEET_MAP = {
    # Domestic Grand Prix — older naming (2021-2024)
    "GP1":  ("PPW", "GP1",  "Grand Prix (runda 1)"),
    "GP2":  ("PPW", "GP2",  "Grand Prix (runda 2)"),
    "GP3":  ("PPW", "GP3",  "Grand Prix (runda 3)"),
    "GP4":  ("PPW", "GP4",  "Grand Prix (runda 4)"),
    "GP5":  ("PPW", "GP5",  "Grand Prix (runda 5)"),
    "GP6":  ("PPW", "GP6",  "Grand Prix (runda 6)"),
    "GP7":  ("PPW", "GP7",  "Grand Prix (runda 7)"),
    "GP8":  ("PPW", "GP8",  "Grand Prix (runda 8)"),
    # Domestic Grand Prix — newer naming (2024-2025)
    "PP1":  ("PPW", "PPW1",  "I Puchar Polski Weteranów"),
    "PP2":  ("PPW", "PPW2",  "II Puchar Polski Weteranów"),
    "PP3":  ("PPW", "PPW3",  "III Puchar Polski Weteranów"),
    "PP4":  ("PPW", "PPW4",  "IV Puchar Polski Weteranów"),
    "PP5":  ("PPW", "PPW5",  "V Puchar Polski Weteranów"),
    # Polish Championships
    "MP":   ("MPW", "MPW",  "Mistrzostwa Polski Weteranów"),  # 2021 naming
    "MPW":  ("MPW", "MPW",  "Mistrzostwa Polski Weteranów"),
    # International EVF Grand Prix — older naming (2022)
    "EVF1": ("PEW", "PEW1", "EVF Grand Prix 1"),
    "EVF2": ("PEW", "PEW2", "EVF Grand Prix 2"),
    "EVF3": ("PEW", "PEW3", "EVF Grand Prix 3"),
    "EVF4": ("PEW", "PEW4", "EVF Grand Prix 4"),
    "EVF5": ("PEW", "PEW5", "EVF Grand Prix 5"),
    "EVF6": ("PEW", "PEW6", "EVF Grand Prix 6"),
    "EVF4'21": ("PEW", "PEW4-21", "EVF Grand Prix 4 (2021)"),
    # International EVF Grand Prix — newer naming (2023+)
    "PEW1": ("PEW", "PEW1", "EVF Grand Prix 1"),
    "PEW2": ("PEW", "PEW2", "EVF Grand Prix 2"),
    "PEW3": ("PEW", "PEW3", "EVF Grand Prix 3"),
    "PEW4": ("PEW", "PEW4", "EVF Grand Prix 4"),
    "PEW5": ("PEW", "PEW5", "EVF Grand Prix 5"),
    "PEW6": ("PEW", "PEW6", "EVF Grand Prix 6"),
    "PEW7": ("PEW", "PEW7", "EVF Grand Prix 7"),
    "PEW8": ("PEW", "PEW8", "EVF Grand Prix 8"),
    "PEW9": ("PEW", "PEW9", "EVF Grand Prix 9"),
    "PEW10":("PEW", "PEW10","EVF Grand Prix 10"),
    "PEW11":("PEW", "PEW11","EVF Grand Prix 11"),
    "PEW12":("PEW", "PEW12","EVF Grand Prix 12"),
    # European/World Championships
    "IMEW": ("MEW", "IMEW", "Indywidualne Mistrzostwa Europy Weteranów"),
    "IMSW": ("MSW", "IMSW", "Indywidualne Mistrzostwa Świata Weteranów"),
    "VFC":  ("MSW", "VFC",  "Veterans Fencing Circuit"),
    "PS":   ("PSW", "PS",   "Puchar Świata"),
}

# Sheets to skip (not events)
SKIP_SHEETS = {"Ranking", "RankingPL", "KlasyfikacjaGP", "KlasyfikacjaEVF", "Kadra"}

# Files to skip (not ranking data)
SKIP_FILE_PATTERNS = {"SuperFive", "czlonkowie", "Backup", "~$"}

# Polish month names for date parsing
POLISH_MONTHS = {
    "stycznia": 1, "lutego": 2, "marca": 3, "kwietnia": 4,
    "maja": 5, "czerwca": 6, "lipca": 7, "sierpnia": 8,
    "września": 9, "października": 10, "listopada": 11, "grudnia": 12,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_date(raw) -> str | None:
    """Return ISO date string from various date formats in the Excel."""
    if raw is None:
        return None
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()

    # "DD.MM.YYYY"
    m = re.match(r"(\d{1,2})\.(\d{2})\.(\d{4})", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"

    # "D-D.MM.YYYY" multi-day event
    m = re.match(r"(\d{1,2})-(\d{1,2})\.(\d{2})\.(\d{4})", s)
    if m:
        d1, d2, mo, y = m.groups()
        return f"{y}-{mo.zfill(2)}-{d1.zfill(2)}"

    # "DD/MM/YYYY"
    m = re.match(r"(\d{1,2})/(\d{2})/(\d{4})", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"

    # "D MONTH_PL YYYY" — e.g. "9 lipca 2022"
    m = re.match(r"(\d{1,2})\s+(\w+)\s+(\d{4})", s)
    if m:
        d, month_name, y = m.groups()
        mo = POLISH_MONTHS.get(month_name.lower())
        if mo:
            return f"{y}-{mo:02d}-{int(d):02d}"

    # bare year
    m = re.match(r"^(\d{4})$", s)
    if m:
        return f"{s}-01-01"

    return None


def parse_date_end(raw) -> str | None:
    """Extract end date from multi-day format like '28-30.09.2024'."""
    if raw is None:
        return None
    s = str(raw).strip()
    m = re.match(r"\d{1,2}-(\d{1,2})\.(\d{2})\.(\d{4})", s)
    if m:
        d2, mo, y = m.groups()
        return f"{y}-{mo.zfill(2)}-{d2.zfill(2)}"
    return None


def derive_event_url(tournament_url: str | None) -> str | None:
    """Derive an event-level URL from a tournament result URL."""
    if not tournament_url:
        return None

    # FTL
    if "fencingtimelive.com" in tournament_url:
        return tournament_url.replace("/events/results/", "/events/")

    # Engarde — strip the last path segment (category) for /competition/ URLs
    if "engarde-service.com" in tournament_url:
        if "/competition/" in tournament_url:
            parts = tournament_url.rstrip("/").rsplit("/", 1)
            if len(parts) == 2:
                return parts[0]
        return tournament_url

    # 4Fence
    if "4fence.it" in tournament_url:
        return re.sub(r"/results.*$", "", tournament_url)

    return tournament_url


def should_skip_file(path: Path) -> bool:
    """Return True if this file should be skipped."""
    full = str(path)
    return any(pat in full for pat in SKIP_FILE_PATTERNS)


def parse_2021_location_date(cell_value) -> tuple[str | None, str | None, str | None]:
    """Parse 2021 format: cell (2,3) contains 'CITY DD-DD.MM.YYYY' combined.

    Returns (location, date_start, date_end).
    """
    if cell_value is None:
        return None, None, None
    s = str(cell_value).strip()

    # If it's a URL, skip
    if s.startswith("http"):
        return None, None, None

    # "CITY DD-DD.MM.YYYY"
    m = re.match(r"(.+?)\s+(\d{1,2})-(\d{1,2})\.(\d{2})\.(\d{4})$", s)
    if m:
        loc, d1, d2, mo, y = m.groups()
        start = f"{y}-{mo}-{int(d1):02d}"
        end = f"{y}-{mo}-{int(d2):02d}"
        return loc.strip(), start, end

    # "CITY DD.MM.YYYY"
    m = re.match(r"(.+?)\s+(\d{1,2})\.(\d{2})\.(\d{4})$", s)
    if m:
        loc, d, mo, y = m.groups()
        dt = f"{y}-{mo}-{int(d):02d}"
        return loc.strip(), dt, dt

    # Just a city name with no date
    if not any(c.isdigit() for c in s):
        return s.strip(), None, None

    return s.strip(), None, None


def detect_layout(wb_data) -> str:
    """Detect whether this workbook uses 2021 or 2022+ layout.

    2021: row 2 has 'miejsce i data:' in (2,1), results start at row 5
    2022+: row 2 has 'miejsce:' in (2,2), row 3 has 'data:', results start at row 6
    """
    # Check any event sheet
    for sn in wb_data.sheetnames:
        if sn in SKIP_SHEETS:
            continue
        if sn not in {k for k in SHEET_MAP}:
            continue
        ws = wb_data[sn]
        val_21 = ws.cell(2, 1).value
        val_22 = ws.cell(2, 2).value
        if val_21 and "miejsce i data" in str(val_21).lower():
            return "2021"
        if val_22 and "miejsce" in str(val_22).lower():
            return "2022+"
        # Fallback: check if row 3 col 2 has ' data:'
        val_32 = ws.cell(3, 2).value
        if val_32 and "data" in str(val_32).lower():
            return "2022+"
        return "2021"  # default for old files
    return "2022+"


def extract_sheet_2022(wb_data, wb_links, sheet_name: str) -> dict:
    """Extract metadata from a 2022+ layout sheet."""
    ws_d = wb_data[sheet_name]
    ws_l = wb_links[sheet_name]

    location = ws_d.cell(2, 3).value
    date_raw = ws_d.cell(3, 3).value
    n = ws_d.cell(2, 8).value

    url = None
    hl = ws_l.cell(2, 3).hyperlink
    if hl:
        url = hl.target

    # Count actual result rows (start at row 6)
    result_count = 0
    for r in range(6, ws_d.max_row + 1):
        name = ws_d.cell(r, 3).value
        place = ws_d.cell(r, 8).value
        if name and str(name).strip() not in ("x", "X", "") and isinstance(place, (int, float)):
            result_count += 1

    return {
        "location": str(location).strip() if location else None,
        "date_start": parse_date(date_raw),
        "date_end": parse_date_end(date_raw),
        "date_raw": str(date_raw).strip() if date_raw else None,
        "participant_count": int(n) if isinstance(n, (int, float)) and n > 0 else result_count,
        "url": url,
        "result_count": result_count,
    }


def extract_sheet_2021(wb_data, wb_links, sheet_name: str) -> dict:
    """Extract metadata from a 2021 layout sheet."""
    ws_d = wb_data[sheet_name]
    ws_l = wb_links[sheet_name]

    cell_23 = ws_d.cell(2, 3).value
    location, date_start, date_end = parse_2021_location_date(cell_23)

    # URL: check hyperlink on (2,5) first, then (2,3)
    url = None
    hl = ws_l.cell(2, 5).hyperlink
    if hl:
        url = hl.target
    else:
        hl = ws_l.cell(2, 3).hyperlink
        if hl:
            url = hl.target
    # Also check if (2,5) has a text URL
    if not url:
        val_25 = ws_d.cell(2, 5).value
        if val_25 and str(val_25).strip().startswith("http"):
            url = str(val_25).strip()

    # Count result rows (start at row 5)
    result_count = 0
    for r in range(5, ws_d.max_row + 1):
        name = ws_d.cell(r, 3).value
        place = ws_d.cell(r, 2).value
        if name and str(name).strip() not in ("x", "X", "") and isinstance(place, (int, float)):
            result_count += 1

    return {
        "location": location,
        "date_start": date_start,
        "date_end": date_end,
        "date_raw": str(cell_23).strip() if cell_23 else None,
        "participant_count": result_count,
        "url": url,
        "result_count": result_count,
    }


# ---------------------------------------------------------------------------
# Season matching
# ---------------------------------------------------------------------------

def load_seasons(conn) -> list[dict]:
    """Load all seasons with date ranges from DB."""
    cur = conn.cursor()
    cur.execute("SELECT id_season, txt_code, dt_start, dt_end FROM tbl_season ORDER BY dt_start")
    seasons = []
    for row in cur.fetchall():
        seasons.append({
            "id_season": row[0],
            "txt_code": row[1],
            "dt_start": row[2],
            "dt_end": row[3],
        })
    return seasons


def find_season_for_date(seasons: list[dict], event_date_str: str | None) -> dict | None:
    """Find the season whose date range contains the event date."""
    if not event_date_str:
        return None
    try:
        event_date = date.fromisoformat(event_date_str)
    except ValueError:
        return None

    for s in seasons:
        if s["dt_start"] <= event_date <= s["dt_end"]:
            return s
    return None


def season_year_suffix(season_code: str) -> str:
    """Convert season code to year suffix for event codes.

    SPWS-2021 → 2021
    SPWS-2023-24 → 2023-2024
    """
    raw = season_code.removeprefix("SPWS-")
    m = re.match(r"(\d{4})-(\d{2})$", raw)
    if m:
        first_year, short = m.groups()
        century = first_year[:2]
        return f"{first_year}-{century}{short}"
    return raw


# ---------------------------------------------------------------------------
# Event extraction
# ---------------------------------------------------------------------------

def extract_events_from_workbook(xlsx_path: Path) -> list[dict]:
    """Extract raw event data from an Excel workbook (no season assignment yet).

    Returns a list of event dicts with event_code_prefix, location, dates, url.
    """
    wb_data = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_links = openpyxl.load_workbook(xlsx_path)

    layout = detect_layout(wb_data)
    extract_fn = extract_sheet_2021 if layout == "2021" else extract_sheet_2022

    events = {}  # keyed by event_code_prefix (e.g. "GP1", "MPW")

    for sheet_name, (ttype, code_prefix, human_name) in SHEET_MAP.items():
        if sheet_name not in wb_data.sheetnames:
            continue

        # Skip EU sheet (EVF rankings, not an event)
        if sheet_name == "EU":
            continue

        data = extract_fn(wb_data, wb_links, sheet_name)

        # Skip sheets with no results
        if data["result_count"] == 0 and data["participant_count"] == 0:
            continue

        organizer = "EVF" if ttype in ("PEW", "MEW", "MSW", "PSW") else "SPWS"
        location = data["location"]
        display_name = f"{human_name} — {location}" if location and ttype in ("PEW", "MEW", "MSW", "PSW") else human_name

        if code_prefix not in events:
            events[code_prefix] = {
                "code_prefix": code_prefix,
                "event_name": display_name,
                "location": location,
                "dt_start": data["date_start"],
                "dt_end": data["date_end"] or data["date_start"],
                "url_event": derive_event_url(data["url"]),
                "organizer": organizer,
                "tournament_type": ttype,
                "source_file": xlsx_path.name,
            }
        else:
            # Merge: expand date range, fill missing URL
            evt = events[code_prefix]
            if data["date_start"]:
                if evt["dt_start"] is None or data["date_start"] < evt["dt_start"]:
                    evt["dt_start"] = data["date_start"]
                end = data["date_end"] or data["date_start"]
                if evt["dt_end"] is None or end > evt["dt_end"]:
                    evt["dt_end"] = end
            if evt["url_event"] is None and data["url"]:
                evt["url_event"] = derive_event_url(data["url"])

    wb_data.close()
    wb_links.close()

    return list(events.values())


def pick_representative_files(xlsx_files: list[Path]) -> list[Path]:
    """Pick one representative file per unique directory (avoid processing
    the same events from every weapon/gender/age file).

    Prefers files containing 'SZPADA-2' or 'SZPADA-0' (epee V2/V0 tend to
    have the most events). Falls back to first file alphabetically.
    """
    by_dir: dict[str, list[Path]] = {}
    for f in xlsx_files:
        key = str(f.parent)
        by_dir.setdefault(key, []).append(f)

    picked = []
    for dir_path, files in sorted(by_dir.items()):
        # Prefer SZPADA-2, then SZPADA-0, then first alphabetically
        best = None
        for f in files:
            name = f.name.upper()
            if "SZPADA-2" in name:
                best = f
                break
            if "SZPADA-0" in name and best is None:
                best = f
        if best is None:
            best = files[0]
        picked.append(best)

    return picked


# ---------------------------------------------------------------------------
# DB operations
# ---------------------------------------------------------------------------

def upsert_events(conn, events_by_season: dict[str, list[dict]], seasons: list[dict],
                  dry_run: bool = False):
    """Insert or update events grouped by season."""
    cur = conn.cursor()

    # Get organizer IDs
    cur.execute("SELECT txt_code, id_organizer FROM tbl_organizer")
    org_map = dict(cur.fetchall())

    # Season code → id mapping
    season_id_map = {s["txt_code"]: s["id_season"] for s in seasons}

    total_created = 0
    total_updated = 0
    total_skipped = 0

    for season_code in sorted(events_by_season.keys()):
        events = events_by_season[season_code]
        season_id = season_id_map.get(season_code)
        if not season_id:
            print(f"\n  ERROR: Season '{season_code}' not found in DB, skipping {len(events)} events")
            continue

        year_suffix = season_year_suffix(season_code)
        print(f"\n{'='*60}")
        print(f"Season: {season_code} (suffix: {year_suffix}) — {len(events)} event(s)")
        print(f"{'='*60}")

        created = 0
        updated = 0
        skipped = 0

        for evt in events:
            event_code = f"{evt['code_prefix']}-{year_suffix}"
            org_id = org_map.get(evt["organizer"])
            if not org_id:
                print(f"  WARN: Organizer '{evt['organizer']}' not found, skipping {event_code}")
                skipped += 1
                continue

            # Check if event already exists
            cur.execute("SELECT id_event, dt_start, dt_end, url_event FROM tbl_event WHERE txt_code = %s",
                        (event_code,))
            existing = cur.fetchone()

            if existing:
                eid, existing_start, existing_end, existing_url = existing
                updates = []
                params = []

                if not existing_start and evt["dt_start"]:
                    updates.append("dt_start = %s")
                    params.append(evt["dt_start"])
                if not existing_end and evt["dt_end"]:
                    updates.append("dt_end = %s")
                    params.append(evt["dt_end"])
                if not existing_url and evt["url_event"]:
                    updates.append("url_event = %s")
                    params.append(evt["url_event"])

                if updates:
                    if dry_run:
                        fields = [u.split(" = ")[0] for u in updates]
                        print(f"  UPDATE {event_code}: filling {', '.join(fields)}")
                    else:
                        params.append(eid)
                        cur.execute(
                            f"UPDATE tbl_event SET {', '.join(updates)} WHERE id_event = %s",
                            params,
                        )
                        fields = [u.split(" = ")[0] for u in updates]
                        print(f"  UPDATE {event_code}: filled {', '.join(fields)}")
                    updated += 1
                else:
                    print(f"  SKIP  {event_code}: already complete")
                    skipped += 1
            else:
                if dry_run:
                    print(f"  INSERT {event_code}: {evt['event_name'][:50]} | {evt['location'] or '?'} | {evt['dt_start']} → {evt['dt_end']}")
                else:
                    cur.execute(
                        """INSERT INTO tbl_event
                           (txt_code, txt_name, txt_location, id_season, id_organizer,
                            dt_start, dt_end, url_event, enum_status)
                           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'COMPLETED')
                           RETURNING id_event""",
                        (
                            event_code,
                            evt["event_name"],
                            evt["location"],
                            season_id,
                            org_id,
                            evt["dt_start"],
                            evt["dt_end"],
                            evt["url_event"],
                        ),
                    )
                    eid = cur.fetchone()[0]
                    print(f"  INSERT {event_code}: id={eid} | {evt['event_name'][:50]} | {evt['location'] or '?'}")
                created += 1

        print(f"  → {created} created, {updated} updated, {skipped} skipped")
        total_created += created
        total_updated += updated
        total_skipped += skipped

    if not dry_run:
        conn.commit()

    print(f"\n{'='*60}")
    print(f"TOTAL: {total_created} created, {total_updated} updated, {total_skipped} skipped")
    return {"created": total_created, "updated": total_updated, "skipped": total_skipped}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Import events from Excel workbooks into the database."
    )
    parser.add_argument("--folder", required=True,
                        help="Root folder to scan recursively for .xlsx files")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview changes without modifying the database")
    parser.add_argument("--json", action="store_true",
                        help="Output results as JSON (for programmatic use)")
    args = parser.parse_args()

    folder = Path(args.folder)
    if not folder.is_dir():
        print(f"ERROR: '{folder}' is not a directory")
        sys.exit(1)

    # Find all .xlsx files recursively, skip non-ranking files
    all_xlsx = sorted(f for f in folder.rglob("*.xlsx") if not should_skip_file(f))
    if not all_xlsx:
        print(f"No .xlsx files found in {folder}")
        sys.exit(1)

    print(f"Found {len(all_xlsx)} Excel file(s) across {folder}")

    # Pick one representative file per directory
    rep_files = pick_representative_files(all_xlsx)
    print(f"Selected {len(rep_files)} representative file(s):")
    for f in rep_files:
        print(f"  {f.relative_to(folder)}")

    if args.dry_run:
        print("\nDRY RUN — no database changes will be made")

    # Connect to DB and load seasons
    conn = psycopg2.connect(DB_URL)
    seasons = load_seasons(conn)
    print(f"\nLoaded {len(seasons)} season(s) from DB:")
    for s in seasons:
        print(f"  {s['txt_code']:20s} {s['dt_start']} → {s['dt_end']}")

    # Extract events from all representative files
    all_raw_events = []
    for xlsx_path in rep_files:
        print(f"\nReading {xlsx_path.relative_to(folder)} ...")
        events = extract_events_from_workbook(xlsx_path)
        all_raw_events.extend(events)
        print(f"  → {len(events)} event(s)")

    if not all_raw_events:
        print("No events found in any Excel file.")
        conn.close()
        sys.exit(0)

    # Assign events to seasons by date, merging duplicates from different weapon files
    # Key: (season_code, code_prefix) → merged event dict
    merged: dict[tuple[str, str], dict] = {}
    no_season = []

    for evt in all_raw_events:
        season = find_season_for_date(seasons, evt["dt_start"])
        if not season:
            no_season.append(evt)
            continue

        season_code = season["txt_code"]
        dedup_key = (season_code, evt["code_prefix"])

        if dedup_key in merged:
            # Merge: extend date range, fill missing fields
            # Guard: skip date extensions > 7 days (likely data entry error)
            existing = merged[dedup_key]
            if evt["dt_start"] and existing["dt_start"]:
                try:
                    new_d = date.fromisoformat(evt["dt_start"])
                    cur_start = date.fromisoformat(existing["dt_start"])
                    cur_end = date.fromisoformat(existing["dt_end"] or existing["dt_start"])
                    # Only extend if new date is within 7 days of current range
                    if new_d < cur_start and (cur_start - new_d).days <= 7:
                        existing["dt_start"] = evt["dt_start"]
                    end_str = evt["dt_end"] or evt["dt_start"]
                    new_end = date.fromisoformat(end_str)
                    if new_end > cur_end and (new_end - cur_end).days <= 7:
                        existing["dt_end"] = end_str
                    elif new_d < cur_start and (cur_start - new_d).days > 7:
                        print(f"  WARN: {evt['code_prefix']} date {evt['dt_start']} differs by >{(cur_start - new_d).days}d from existing {existing['dt_start']}, skipping")
                except ValueError:
                    pass
            elif evt["dt_start"] and existing["dt_start"] is None:
                existing["dt_start"] = evt["dt_start"]
                existing["dt_end"] = evt["dt_end"] or evt["dt_start"]
            if existing["url_event"] is None and evt.get("url_event"):
                existing["url_event"] = evt["url_event"]
            if existing["location"] is None and evt.get("location"):
                existing["location"] = evt["location"]
        else:
            merged[dedup_key] = dict(evt)

    # Group merged events by season
    events_by_season: dict[str, list[dict]] = {}
    for (season_code, _), evt in merged.items():
        events_by_season.setdefault(season_code, []).append(evt)

    # Print summary before upsert
    print(f"\n{'='*60}")
    print(f"Events by season:")
    for sc in sorted(events_by_season.keys()):
        evts = events_by_season[sc]
        suffix = season_year_suffix(sc)
        for e in evts:
            code = f"{e['code_prefix']}-{suffix}"
            print(f"  {sc:20s} {code:25s} {e['location'] or '?':20s} {e['dt_start'] or '?':12s}")

    if no_season:
        print(f"\nWARNING: {len(no_season)} event(s) could not be assigned to any season:")
        for evt in no_season:
            print(f"  {evt['code_prefix']:10s} {evt['location'] or '?':20s} date={evt['dt_start']}")

    # Upsert
    try:
        result = upsert_events(conn, events_by_season, seasons, dry_run=args.dry_run)
        conn.close()
    except Exception as e:
        print(f"ERROR: {e}")
        conn.rollback()
        conn.close()
        sys.exit(1)

    if args.json:
        print(json.dumps(result))


if __name__ == "__main__":
    main()
