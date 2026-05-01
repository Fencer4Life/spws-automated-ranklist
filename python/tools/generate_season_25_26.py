"""
generate_season_25_26.py
------------------------
Reads doc/assets/SZPADA-2-2025-2026.xlsx and:
  1. Reports new fencers found in the Ranking tab (for manual review / seed.sql update)
  2. Generates supabase/data/season_2025_26.sql with tournament + result data

Run: python python/tools/generate_season_25_26.py

Structure differences from 2024-2025 Excel:
  - Some sheets have a 'LINK' header row, some don't
  - Location/N detected by scanning for 'miejsce:' / 'turniej:' in col B
  - Data rows always start at row 6; name=col3, place=col8 (same as before)
  - URLs may appear as literal text in col C (not just hyperlinks)
"""

import re, json, sys
from pathlib import Path
from datetime import date, datetime

import openpyxl
import psycopg2
from rapidfuzz import fuzz

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
XLSX_PATH = Path("doc/assets/SZPADA-2-2025-2026.xlsx")
OUT_PATH   = Path("supabase/data/season_2025_26.sql")
SEASON_CODE = "SPWS-2025-2026"
DB_URL = "postgresql://postgres:postgres@127.0.0.1:54322/postgres"

# Map sheet name → (tournament_type, code_prefix, human_name)
SHEET_MAP = {
    "PP1":  ("PPW", "PP1",  "I Puchar Polski Weteranów — Szpada M"),
    "PP2":  ("PPW", "PP2",  "II Puchar Polski Weteranów — Szpada M"),
    "PP3":  ("PPW", "PP3",  "III Puchar Polski Weteranów — Szpada M"),
    "PP4":  ("PPW", "PP4",  "IV Puchar Polski Weteranów — Szpada M"),
    "PP5":  ("PPW", "PP5",  "V Puchar Polski Weteranów — Szpada M"),
    "MPW":  ("MPW", "MPW",  "Mistrzostwa Polski Weteranów — Szpada M"),
    "PEW1": ("PEW", "PEW1", "EVF Grand Prix 1 — Budapeszt"),
    "PEW2": ("PEW", "PEW2", "EVF Grand Prix 2 — Madryt"),
    "PEW3": ("PEW", "PEW3", "EVF Grand Prix 3 — Guildford"),
    "PEW4": ("PEW", "PEW4", "EVF Grand Prix 4 — Terni"),
    "PEW5": ("PEW", "PEW5", "EVF Grand Prix 5 — Sztokholm"),
    "PEW6": ("PEW", "PEW6", "EVF Grand Prix 6 — Warszawa"),
    "PEW7": ("PEW", "PEW7", "EVF Grand Prix 7 — Chania"),
    "PS":   ("PEW", "PS",   "Puchar Świata Weteranów — Paryż"),
    "IMEW": ("MEW", "IMEW", "Indywidualne Mistrzostwa Europy Weteranów — Płowdiw"),
    "IMSW": ("MSW", "IMSW", "Indywidualne Mistrzostwa Świata Weteranów"),
}

MATCH_THRESHOLD = 80

SPANISH_MONTHS = {
    "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
    "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
    "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sq(s):
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def parse_date(raw) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()

    # "DD.MM.YYYY"
    m = re.match(r"(\d{1,2})\.(\d{2})\.(\d{4})", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"

    # "DD-DD.MM.YYYY" multi-day
    m = re.match(r"(\d{1,2})-\d{1,2}\.(\d{2})\.(\d{4})", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"

    # "YYYY.MM.DD" (Budapest style)
    m = re.match(r"(\d{4})\.(\d{2})\.(\d{2})", s)
    if m:
        y, mo, d = m.groups()
        return f"{y}-{mo}-{d}"

    # "D - MONTH_ES - YYYY" (Spanish)
    m = re.match(r"(\d{1,2})\s*-\s*(\w+)\s*-(\d{4})", s, re.IGNORECASE)
    if m:
        d, mon, y = m.groups()
        mo = SPANISH_MONTHS.get(mon.lower())
        if mo:
            return f"{y}-{mo}-{d.zfill(2)}"

    # "Weekday, Month DD, YYYY H:MM AM/PM" (FencingTimeLive style)
    m = re.match(r"\w+,\s+(\w+)\s+(\d+),\s+(\d{4})", s)
    if m:
        month_name, d, y = m.groups()
        months = {
            "January":"01","February":"02","March":"03","April":"04",
            "May":"05","June":"06","July":"07","August":"08",
            "September":"09","October":"10","November":"11","December":"12",
        }
        mo = months.get(month_name)
        if mo:
            return f"{y}-{mo}-{d.zfill(2)}"

    # bare year
    m = re.match(r"^(\d{4})$", s)
    if m:
        return f"{s}-01-01"

    return None


def normalize_name(name: str) -> str:
    return " ".join(name.upper().split())


def fuzzy_match(name: str, fencers: list[tuple]) -> tuple | None:
    norm = normalize_name(name)
    for fid, surname, first, aliases in fencers:
        for alias in aliases:
            if normalize_name(alias) == norm:
                return (fid, f"{surname} {first}", 100)
    best_score = 0
    best_fencer = None
    for fid, surname, first, aliases in fencers:
        full = f"{surname} {first}"
        score = fuzz.token_sort_ratio(norm, full.upper())
        if score > best_score:
            best_score = score
            best_fencer = (fid, f"{surname} {first}")
    if best_score >= MATCH_THRESHOLD:
        return (*best_fencer, best_score)
    return None


def load_fencers(conn) -> list[tuple]:
    cur = conn.cursor()
    cur.execute("""
        SELECT id_fencer, txt_surname, txt_first_name,
               COALESCE(json_name_aliases::text, '[]')
          FROM tbl_fencer ORDER BY id_fencer
    """)
    return [(r[0], r[1], r[2], json.loads(r[3])) for r in cur.fetchall()]


# ---------------------------------------------------------------------------
# Extract tournament metadata + results from one sheet
# ---------------------------------------------------------------------------

def extract_sheet(wb_data, wb_links, sheet_name: str) -> dict:
    ws_d = wb_data[sheet_name]
    ws_l = wb_links[sheet_name]

    location = None
    date_raw = None
    n = None
    url = None

    # Scan rows 1-5 for metadata
    for r in range(1, 6):
        col_b = str(ws_d.cell(r, 2).value or "").strip().lower()
        col_c = ws_d.cell(r, 3).value
        col_h = ws_d.cell(r, 8).value

        if "miejsce" in col_b:
            if location is None and col_c:
                location = str(col_c).strip()
            if n is None and isinstance(col_h, (int, float)) and float(col_h) > 0:
                n = int(col_h)

        if "turniej" in col_b:
            # LINK-format: N is in this row
            if n is None and isinstance(col_h, (int, float)) and float(col_h) > 0:
                n = int(col_h)

        if "data" in col_b and date_raw is None:
            date_raw = col_c

        # URL as literal text in col C
        if col_c and isinstance(col_c, str) and col_c.strip().lower().startswith("http"):
            url = col_c.strip()
        # URL as hyperlink
        if url is None:
            hl = ws_l.cell(r, 3).hyperlink
            if hl:
                url = hl.target

    # Data rows: row 6 onwards
    results = []
    for r in range(6, ws_d.max_row + 1):
        name  = ws_d.cell(r, 3).value
        place = ws_d.cell(r, 8).value
        if not name or str(name).strip().upper() in ("X", ""):
            continue
        if not isinstance(place, int):
            continue
        results.append({"name": str(name).strip(), "place": place})

    return {
        "location": location,
        "date":     parse_date(date_raw),
        "n":        n,
        "url":      url,
        "results":  results,
    }


# ---------------------------------------------------------------------------
# Ranking tab: extract fencer names
# ---------------------------------------------------------------------------

def extract_ranking_names(wb_data) -> list[tuple[int, str]]:
    ws = wb_data["Ranking"]
    names = []
    for r in range(6, ws.max_row + 1):
        rank = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if rank is None or not isinstance(rank, (int, float)):
            continue
        if not name:
            continue
        names.append((int(rank), str(name).strip()))
    return names


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"Reading {XLSX_PATH} ...")
    wb_data  = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    wb_links = openpyxl.load_workbook(XLSX_PATH)

    print("Connecting to DB ...")
    conn = psycopg2.connect(DB_URL)
    fencers = load_fencers(conn)
    print(f"  Loaded {len(fencers)} fencers from DB")
    conn.close()

    # --- Check Ranking tab for new fencers ---
    ranking_names = extract_ranking_names(wb_data)
    print(f"\nRanking tab: {len(ranking_names)} fencers")
    new_fencers = []
    for rank, name in ranking_names:
        match = fuzzy_match(name, fencers)
        if match:
            print(f"  ✓ [{match[2]:3.0f}] {rank:2d}. {name} → {match[1]}")
        else:
            print(f"  NEW       {rank:2d}. {name}")
            new_fencers.append(name)

    if new_fencers:
        print(f"\n*** {len(new_fencers)} NEW fencers to add to seed.sql:")
        for name in new_fencers:
            parts = name.split(None, 1)
            surname = parts[0] if parts else name
            first   = parts[1] if len(parts) > 1 else ""
            print(f"    INSERT INTO tbl_fencer (txt_surname, txt_first_name) VALUES ('{surname}', '{first}');")

    # --- Generate SQL ---
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "-- =========================================================================",
        f"-- Season 2025/26 real data — generated from {XLSX_PATH.name}",
        "-- Run AFTER seed.sql (which creates the season and organizers).",
        "-- =========================================================================",
        "",
        "-- Expand season start to cover EVF rounds from early 2025",
        "UPDATE tbl_season",
        "   SET dt_start = '2025-01-01'",
        f" WHERE txt_code = '{SEASON_CODE}';",
        "",
        "-- Remove the placeholder sample event/tournament from seed.sql",
        "DELETE FROM tbl_tournament WHERE txt_code = 'PPW1-V2-M-EPEE-2025';",
        "DELETE FROM tbl_event WHERE txt_code = 'PPW1-KRAKOW-2025';",
        "",
    ]

    total_matched = 0
    total_unmatched = 0

    for sheet_name, (ttype, code, human_name) in SHEET_MAP.items():
        if sheet_name not in wb_data.sheetnames:
            print(f"  SKIP {sheet_name}: not in workbook")
            continue

        data = extract_sheet(wb_data, wb_links, sheet_name)
        loc  = data["location"] or "?"
        dt   = data["date"]
        n    = data["n"]
        url  = data["url"]

        if not n:
            print(f"  SKIP {sheet_name}: N={n} (no participants)")
            lines += [f"-- SKIP {sheet_name} ({human_name}): N={n} — tournament skipped", ""]
            continue

        event_code = f"{code}-2025-2026"
        tourn_code = f"{code}-V2-M-EPEE-2025-2026"

        organizer_raw = "EVF" if ttype in ("PEW", "MEW", "MSW") else "SPWS"

        lines += [
            f"-- ---- {sheet_name}: {human_name} ({loc}) ----",
            f"INSERT INTO tbl_event (txt_code, txt_name, id_season, id_organizer, enum_status)",
            f"VALUES (",
            f"    {sq(event_code)},",
            f"    {sq(human_name)},",
            f"    (SELECT id_season FROM tbl_season WHERE txt_code = {sq(SEASON_CODE)}),",
            f"    (SELECT id_organizer FROM tbl_organizer WHERE txt_code = {sq(organizer_raw)}),",
            f"    'COMPLETED'",
            f");",
            f"INSERT INTO tbl_tournament (",
            f"    id_event, txt_code, txt_name, enum_type,",
            f"    enum_weapon, enum_gender, enum_age_category,",
            f"    dt_tournament, int_participant_count, url_results,",
            f"    enum_import_status",
            f") VALUES (",
            f"    (SELECT id_event FROM tbl_event WHERE txt_code = {sq(event_code)}),",
            f"    {sq(tourn_code)},",
            f"    {sq(human_name)},",
            f"    '{ttype}',",
            f"    'EPEE', 'M', 'V2',",
            f"    {sq(dt)}, {n if n else 'NULL'}, {sq(url)},",
            f"    'SCORED'",
            f");",
        ]

        matched = 0
        unmatched = 0
        for row in data["results"]:
            match = fuzzy_match(row["name"], fencers)
            if match:
                fid, matched_name, score = match
                matched += 1
                lines += [
                    f"INSERT INTO tbl_result (id_fencer, id_tournament, int_place, txt_scraped_name)",
                    f"VALUES (",
                    f"    {fid},",
                    f"    (SELECT id_tournament FROM tbl_tournament WHERE txt_code = {sq(tourn_code)}),",
                    f"    {row['place']},",
                    f"    {sq(row['name'])}",
                    f"); -- matched: {matched_name} (score={score:.0f})",
                ]
            else:
                unmatched += 1
                lines += [f"-- UNMATCHED (<{MATCH_THRESHOLD}): {sq(row['name'])} place={row['place']}"]

        total_matched += matched
        total_unmatched += unmatched
        print(f"  {sheet_name}: N={n}, results={len(data['results'])}, matched={matched}, unmatched={unmatched}")

        lines += [
            f"-- Compute scores for {tourn_code}",
            f"SELECT fn_calc_tournament_scores(",
            f"    (SELECT id_tournament FROM tbl_tournament WHERE txt_code = {sq(tourn_code)})",
            f");",
            "",
        ]

    lines += [
        "-- Summary",
        f"-- Total results matched:   {total_matched}",
        f"-- Total results unmatched: {total_unmatched}",
        "",
    ]

    OUT_PATH.write_text("\n".join(lines), encoding="utf-8")
    print(f"\nWrote {OUT_PATH}")
    print(f"Total: matched={total_matched}, unmatched={total_unmatched}")


if __name__ == "__main__":
    main()
