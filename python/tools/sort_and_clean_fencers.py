#!/usr/bin/env python3
"""Sort seed_tbl_fencer.sql alphabetically, remove duplicates and non-domestic fencers, remap all IDs.

This script:
1. Removes specified duplicate fencer entries
2. Removes fencers with no PPW/MPW results (international-only or zero-result) — ADR-019
3. Sorts all remaining fencer entries alphabetically by (surname, first_name)
4. Remaps ALL fencer IDs in supabase/data/**/*.sql
5. Deletes result INSERT blocks for removed fencers from data files
6. Adds alias UPDATEs for identity resolution

Usage:
    python python/tools/sort_and_clean_fencers.py --dry-run   # preview
    python python/tools/sort_and_clean_fencers.py             # apply
"""

import locale
import re
import sys
from pathlib import Path

import openpyxl
from rapidfuzz import fuzz

# Add project python dir to path for imports
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from matcher.fuzzy_match import normalize_name, fold_diacritics

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
SEED_PATH = PROJECT_ROOT / "supabase" / "seed_tbl_fencer.sql"
DATA_DIR = PROJECT_ROOT / "supabase" / "data"
EXCEL_DIR = PROJECT_ROOT / "doc" / "external_files"

# Seasons that only exist as Excel (no SQL data files)
EXCEL_ONLY_SEASONS = ["Sezon 2021", "Sezon 2022"]

# Duplicates to remove: {old_id: primary_id}
# Run 1 (already applied): KORONA-TRZEBSKI→KORONA, TK→KOŃCZYŁO
# Run 2 (already applied): 6 duplicate pairs
DUPLICATES = {
    66: 65,    # FRAŚ Felix → FRAŚ Feliks
    72: 71,    # FUHRMANN Urlike → FUHRMANN Ulrike
    170: 169,  # KRUJASKIS Gotfridas → KRUJALSKIS Gotfridas
    196: 195,  # MAZIK Alksander → MAZIK Aleksander
    216: 217,  # NIKALAICHUK Aleksander → NIKALAICHUK Aliaksandr
    348: 349,  # WOJTAS Bogdan → WOJTAS Bogusław
}

# Aliases to write at the end of the seed file
# (surname, first_name) → list of aliases
ALIASES = {
    ("KOŃCZYŁO", "Tomasz"): ["TK"],
    ("SZMAJDZIŃSKA", "Katarzyna"): ["SZMAJDZIŃSKA - BOŁDYS Katarzyna"],
    ("KORONA", "Przemysław"): ["KORONA-TRZEBSKI Przemysław"],
    ("FRAŚ", "Feliks"): ["FRAŚ Felix"],
    ("FUHRMANN", "Ulrike"): ["FUHRMANN Urlike"],
    ("KRUJALSKIS", "Gotfridas"): ["KRUJASKIS Gotfridas"],
    ("MAZIK", "Aleksander"): ["MAZIK Alksander"],
    ("NIKALAICHUK", "Aliaksandr"): ["NIKALAICHUK Aleksander"],
    ("WOJTAS", "Bogusław"): ["WOJTAS Bogdan"],
}


def scan_ppw_mpw_fencer_ids(data_dir: Path) -> set[int]:
    """Scan all data SQL files and return fencer IDs that appear in PPW/MPW/GPW results.

    Derives the domestic fencer set dynamically from actual data files,
    so the set never goes stale (ADR-019).
    """
    id_line_pattern = re.compile(r"^\s+(\d+),\s*$")
    ppw_mpw_ids = set()

    for sql_file in sorted(data_dir.rglob("*.sql")):
        if sql_file.name.startswith("zz_") or sql_file.name == "season_config.sql":
            continue

        lines = sql_file.read_text(encoding="utf-8").splitlines()
        for i, line in enumerate(lines):
            if i >= 2 and "tbl_result" in lines[i - 2]:
                m = id_line_pattern.match(line)
                if m:
                    fid = int(m.group(1))
                    # Check tournament code in the SELECT line (i+1 or i+2)
                    for j in range(i + 1, min(i + 3, len(lines))):
                        if "txt_code" in lines[j]:
                            code_line = lines[j]
                            # PPW/MPW/GPW are domestic; PEW/MEW/PSW/MSW/IMEW/IMSW are international
                            is_domestic = any(
                                t in code_line
                                for t in ["PPW", "MPW", "GPW"]
                            )
                            is_international = any(
                                t in code_line
                                for t in ["PEW", "MEW", "PSW", "MSW", "IMEW", "IMSW"]
                            )
                            if is_domestic and not is_international:
                                ppw_mpw_ids.add(fid)
                            break

    return ppw_mpw_ids


def scan_excel_domestic_names(excel_dir: Path, season_dirs: list[str]) -> set[str]:
    """Extract fencer names from Excel-only seasons (domestic GP/MP sheets).

    Returns normalized names (uppercase "SURNAME FirstName") for fuzzy matching
    against the seed. Skips EU/Ranking/SuperFive sheets (international/summary).
    """
    names = set()
    for season_name in season_dirs:
        season_path = excel_dir / season_name
        if not season_path.exists():
            continue
        for xlsx_path in sorted(season_path.rglob("*.xlsx")):
            if "SuperFive" in xlsx_path.name or "Backup" in str(xlsx_path):
                continue
            try:
                wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
            except Exception:
                continue
            for sheet_name in wb.sheetnames:
                if sheet_name.upper() in ("EU", "RANKING"):
                    continue
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    if len(row) >= 3 and row[2] and isinstance(row[2], str):
                        name = row[2].strip()
                        if name in ("Name", "IMIĘ NAZWISKO", "Nazwisko Imię") or len(name) < 4:
                            continue
                        # Filter out dates/cities/URLs
                        if any(c.isdigit() for c in name) and "." in name:
                            continue
                        if "http" in name.lower():
                            continue
                        if " " in name:
                            names.add(name)
            wb.close()
    return names


def match_excel_names_to_seed(
    excel_names: set[str], entries: list[dict], threshold: float = 90.0
) -> set[int]:
    """Fuzzy-match Excel fencer names against seed entries.

    Returns set of seed old_ids that match an Excel name at ≥threshold.
    Uses diacritic folding + token_sort_ratio for cross-source matching.
    """
    matched_ids = set()
    for excel_name in excel_names:
        excel_norm = normalize_name(excel_name)
        excel_folded = fold_diacritics(excel_norm)

        best_score = 0.0
        best_entry = None
        for entry in entries:
            seed_name = f"{entry['surname']} {entry['first_name']}"
            seed_norm = normalize_name(seed_name)
            seed_folded = fold_diacritics(seed_norm)
            score = max(
                fuzz.token_sort_ratio(excel_norm, seed_norm),
                fuzz.token_sort_ratio(excel_folded, seed_folded),
            )
            if score > best_score:
                best_score = score
                best_entry = entry

        if best_score >= threshold and best_entry is not None:
            matched_ids.add(best_entry["old_id"])

    return matched_ids


def parse_seed_entries(seed_path: Path) -> list[dict]:
    """Parse fencer entries from seed file. Returns list of dicts with id, surname, first_name, etc."""
    text = seed_path.read_text(encoding="utf-8")
    lines = text.splitlines()

    entries = []
    # Each data line: ('SURNAME', 'FirstName', YEAR_OR_NULL),  -- optional comment
    pattern = re.compile(
        r"\s*\('([^']*)',\s*'([^']*)',\s*(NULL|\d+)\s*\)\s*([,;])\s*(--\s*.*)?"
    )

    # Auto-increment ID based on order of appearance (not line number)
    fencer_id = 0
    for i, line in enumerate(lines):
        m = pattern.match(line)
        if m:
            fencer_id += 1
            surname, first_name, year_str, _sep, comment = m.groups()
            # Strip NO PPW/MPW comment if present
            if comment and "NO PPW/MPW" in comment:
                comment = None
            entries.append({
                "old_id": fencer_id,
                "surname": surname,
                "first_name": first_name,
                "birth_year": None if year_str == "NULL" else int(year_str),
                "comment": comment.strip() if comment else None,
                "raw_line": line,
            })

    return entries


def sort_key(entry: dict) -> tuple:
    """Sort by surname then first_name using Polish locale collation."""
    return (
        locale.strxfrm(entry["surname"].upper()),
        locale.strxfrm(entry["first_name"].upper()),
    )


def build_fencer_line(entry: dict, is_last: bool) -> str:
    """Build a formatted seed line for a fencer entry."""
    surname = entry["surname"]
    first_name = entry["first_name"]
    year = "NULL" if entry["birth_year"] is None else str(entry["birth_year"])

    # Pad for alignment
    surname_field = f"'{surname}',"
    name_field = f"'{first_name}',"
    # Align columns
    surname_padded = surname_field.ljust(30)
    name_padded = name_field.ljust(22)

    sep = ";" if is_last else ","
    base = f"    ({surname_padded} {name_padded} {year}){sep}"

    if entry.get("comment"):
        comment_text = entry["comment"].lstrip("- ").strip()
        return f"{base} -- {comment_text}"
    return base


def remap_data_files(
    data_dir: Path,
    id_map: dict[int, int],
    removed_ids: set[int],
    dry_run: bool,
) -> tuple[int, int]:
    """Update fencer ID references and delete result blocks for removed fencers.

    id_map: old_id → new_id for kept fencers
    removed_ids: old IDs whose result blocks should be deleted

    Returns (remap_count, delete_count).
    """
    id_line_pattern = re.compile(r"^(\s+)(\d+)(,\s*)$")
    total_remapped = 0
    total_deleted = 0

    for sql_file in sorted(data_dir.rglob("*.sql")):
        if sql_file.name.startswith("zz_") or sql_file.name == "season_config.sql":
            continue

        lines = sql_file.read_text(encoding="utf-8").splitlines(keepends=True)
        lines_to_delete = set()
        remap_changes = {}  # line_index → new_line

        # Pass 1: identify lines to delete and lines to remap
        for i, line in enumerate(lines):
            if i >= 2 and "tbl_result" in lines[i - 2]:
                m = id_line_pattern.match(line)
                if m:
                    indent, id_str, suffix = m.groups()
                    old_id = int(id_str)

                    if old_id in removed_ids:
                        # Delete the entire result INSERT block (7 lines):
                        # i-2: INSERT INTO tbl_result ...
                        # i-1: VALUES (
                        # i:       {ID},
                        # i+1:     (SELECT id_tournament ...),
                        # i+2:     {place},
                        # i+3:     {name}
                        # i+4: ); -- matched: ...
                        for j in range(i - 2, min(i + 5, len(lines))):
                            lines_to_delete.add(j)
                        total_deleted += 1
                        if dry_run:
                            print(f"  DELETE {sql_file.name}:{i+1}  id {old_id}")
                    elif old_id in id_map and id_map[old_id] != old_id:
                        new_id = id_map[old_id]
                        remap_changes[i] = f"{indent}{new_id}{suffix}"
                        total_remapped += 1
                        if dry_run:
                            print(f"  REMAP  {sql_file.name}:{i+1}  id {old_id} → {new_id}")

        if not lines_to_delete and not remap_changes:
            continue

        # Pass 2: build new file content
        new_lines = []
        for i, line in enumerate(lines):
            if i in lines_to_delete:
                continue
            elif i in remap_changes:
                new_lines.append(remap_changes[i])
            else:
                new_lines.append(line)

        if not dry_run:
            sql_file.write_text("".join(new_lines), encoding="utf-8")

    return total_remapped, total_deleted


def write_seed_file(seed_path: Path, entries: list[dict]):
    """Write the sorted seed file with header, entries, and alias UPDATEs."""
    count = len(entries)
    header = f"""\
-- =============================================================================
-- Master Fencer List — tbl_fencer
-- {count} SPWS members (domestic PPW/MPW participants only — ADR-019).
-- Birth year only; club and nationality not tracked.
-- Auto-loaded via config.toml sql_paths glob after seed.sql.
-- Note: birth year alone is sufficient for SPWS age-category rules (calendar-year-based).
-- NULL int_birth_year = year unknown.
-- =============================================================================
INSERT INTO tbl_fencer (txt_surname, txt_first_name, int_birth_year) VALUES
"""

    lines = [header]
    for i, entry in enumerate(entries):
        is_last = i == len(entries) - 1
        lines.append(build_fencer_line(entry, is_last))

    # Add alias UPDATEs
    lines.append("")
    lines.append("-- Name aliases for identity resolution (M4/M5)")

    for (surname, first_name), aliases in sorted(ALIASES.items()):
        alias_json = "[" + ", ".join(f'"{a}"' for a in aliases) + "]"
        lines.append(
            f"-- {surname} {first_name} also competed as: {', '.join(aliases)}"
        )
        lines.append(
            f"UPDATE tbl_fencer SET json_name_aliases = '{alias_json}'"
        )
        lines.append(
            f"WHERE txt_surname = '{surname}' AND txt_first_name = '{first_name}';"
        )

    seed_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    dry_run = "--dry-run" in sys.argv

    # Set Polish locale for correct alphabetical sorting
    try:
        locale.setlocale(locale.LC_ALL, "pl_PL.UTF-8")
    except locale.Error:
        print("WARNING: pl_PL.UTF-8 locale not available, falling back to default")
        locale.setlocale(locale.LC_ALL, "")

    if dry_run:
        print("=== DRY RUN ===\n")
    else:
        print("=== APPLYING CHANGES ===\n")

    # Step 1: Parse current entries
    entries = parse_seed_entries(SEED_PATH)
    print(f"Parsed {len(entries)} fencer entries")

    # Step 2: Remove duplicates
    kept = []
    removed_names = []
    for e in entries:
        if e["old_id"] in DUPLICATES:
            removed_names.append(f"  id={e['old_id']}: {e['surname']} {e['first_name']}")
        else:
            kept.append(e)
    if removed_names:
        print(f"Removed {len(removed_names)} duplicates:")
        for name in removed_names:
            print(name)
    else:
        print("No duplicates to remove")

    # Step 3: Scan data files to find domestic (PPW/MPW) fencer IDs
    print(f"\nScanning SQL data files for PPW/MPW participation...")
    ppw_mpw_ids = scan_ppw_mpw_fencer_ids(DATA_DIR)
    print(f"  Found {len(ppw_mpw_ids)} fencers with PPW/MPW results in SQL")

    # Step 3b: Scan Excel-only seasons (2021, 2022) for domestic participants
    print(f"Scanning Excel-only seasons for domestic participants...")
    excel_names = scan_excel_domestic_names(EXCEL_DIR, EXCEL_ONLY_SEASONS)
    print(f"  Found {len(excel_names)} unique names in Excel files")
    excel_ids = match_excel_names_to_seed(excel_names, kept)
    excel_only_ids = excel_ids - ppw_mpw_ids
    print(f"  Matched {len(excel_ids)} to seed ({len(excel_only_ids)} Excel-only, not in SQL)")
    domestic_ids = ppw_mpw_ids | excel_ids

    # Step 4: Remove non-domestic fencers (ADR-019)
    before_count = len(kept)
    non_domestic = []
    domestic = []
    for e in kept:
        if e["old_id"] in domestic_ids:
            domestic.append(e)
        else:
            non_domestic.append(e)
    kept = domestic
    print(f"Removed {len(non_domestic)} non-domestic fencers (no PPW/MPW results)")
    if dry_run and non_domestic:
        for e in non_domestic[:10]:
            print(f"  id={e['old_id']}: {e['surname']} {e['first_name']}")
        if len(non_domestic) > 10:
            print(f"  ... and {len(non_domestic) - 10} more")

    # Step 5: Sort alphabetically
    kept.sort(key=sort_key)
    print(f"\nSorted {len(kept)} entries alphabetically")

    # Step 6: Build complete ID mapping (old_id → new_id)
    id_map = {}
    # First, map duplicates to their primaries
    for dup_id, primary_id in DUPLICATES.items():
        id_map[dup_id] = primary_id  # temporary — will be updated after sorting

    # Map old_id → new_id for all kept entries
    old_to_new = {}
    for new_id_0based, entry in enumerate(kept):
        new_id = new_id_0based + 1
        old_to_new[entry["old_id"]] = new_id

    # Now update duplicate mappings to use the primary's NEW id
    for dup_id, primary_id in DUPLICATES.items():
        if primary_id in old_to_new:
            id_map[dup_id] = old_to_new[primary_id]
        # If primary was also removed (non-domestic), skip silently

    # Merge kept entries' mapping
    id_map.update(old_to_new)

    changes = sum(1 for old, new in id_map.items() if old != new)
    print(f"ID mappings: {len(id_map)} total, {changes} changed")

    # Compute removed IDs (duplicates + non-domestic) for result block deletion
    all_seed_ids = {e["old_id"] for e in entries}
    kept_ids = {e["old_id"] for e in kept}
    removed_ids = all_seed_ids - kept_ids
    print(f"IDs to remove from data files: {len(removed_ids)}")

    # Step 7: Remap data files + delete removed result blocks
    print(f"\nProcessing data files...")
    remapped, deleted = remap_data_files(DATA_DIR, id_map, removed_ids, dry_run)
    print(f"  Remapped: {remapped}")
    print(f"  Deleted result blocks: {deleted}")

    # Print ID mapping for key fencers (useful for pgTAP test updates)
    key_fencers = {}
    for entry in kept:
        key_fencers[entry["old_id"]] = (
            old_to_new[entry["old_id"]],
            entry["surname"],
            entry["first_name"],
        )
    print(f"\nKey ID mappings for pgTAP test updates:")
    # Print mappings for known test-referenced fencers
    test_ids = [9, 56, 105, 106, 117, 145, 228, 235, 317, 318, 331, 335, 355]
    for old_id in test_ids:
        if old_id in key_fencers:
            new_id, surname, first = key_fencers[old_id]
            print(f"  {old_id} → {new_id}  ({surname} {first})")
        else:
            print(f"  {old_id} → REMOVED")

    # Step 8: Write seed file
    if not dry_run:
        print(f"\nWriting sorted seed file ({len(kept)} fencers)...")
        write_seed_file(SEED_PATH, kept)
    else:
        print(f"\nWould write sorted seed file ({len(kept)} fencers)")
        # Show first/last few entries
        print("  First 5:")
        for e in kept[:5]:
            print(f"    {e['surname']} {e['first_name']} (was id={e['old_id']})")
        print("  Last 5:")
        for e in kept[-5:]:
            print(f"    {e['surname']} {e['first_name']} (was id={e['old_id']})")

    print(f"\nDone!")
    if dry_run:
        print("Re-run without --dry-run to apply.")


if __name__ == "__main__":
    main()
